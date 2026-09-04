from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook


_FORMULA_SAFE = True  # cells are written as values, not formulas

RESERVED_ENTITY = frozenset({"name", "entity_type", "org_unit"})
RESERVED_INSTANCE = frozenset({"entity", "form", "org_unit"})


def xlsform_version_today() -> str:
    return datetime.now().strftime("%Y%m%d") + "01"


def infer_field_type(name: str, values: list[str]) -> str:
    lowered = name.lower()
    if lowered in {"sex", "gender"}:
        return "select_one sex"
    if "date" in lowered:
        return "date"
    if lowered in {"age", "muac", "muac_mm"}:
        return "integer"
    if "weight" in lowered:
        return "decimal"
    sample = [value for value in values if value]
    if sample and all(value.replace(".", "", 1).isdigit() for value in sample):
        if any("." in value for value in sample):
            return "decimal"
        return "integer"
    return "text"


def infer_fields(rows: list[dict[str, str]], reserved: frozenset[str]) -> list[dict[str, str]]:
    names: list[str] = []
    for row in rows:
        for key in row:
            if key not in reserved and key not in names and key:
                names.append(key)
    fields: list[dict[str, str]] = []
    for name in names:
        values = [str(row.get(name) or "") for row in rows]
        fields.append(
            {
                "name": name,
                "type": infer_field_type(name, values),
                "label": name.replace("_", " ").title(),
            }
        )
    return fields


def build_xlsform_bytes(
    *,
    form_title: str,
    form_id: str,
    version: str,
    fields: list[dict[str, str]],
) -> bytes:
    workbook = Workbook()
    survey = workbook.active
    survey.title = "survey"
    survey.append(["type", "name", "label", "required"])
    survey.append(["start", "start", "Start", ""])
    survey.append(["end", "end", "End", ""])
    needs_sex = False
    for field in fields:
        required = "yes" if field["name"] in {"first_name", "last_name"} else ""
        survey.append([field["type"], field["name"], field["label"], required])
        if field["type"].startswith("select_one sex"):
            needs_sex = True

    choices = workbook.create_sheet("choices")
    choices.append(["list_name", "name", "label"])
    if needs_sex:
        choices.append(["sex", "F", "Female"])
        choices.append(["sex", "M", "Male"])

    settings = workbook.create_sheet("settings")
    settings.append(["form_title", "form_id", "version"])
    settings.append([form_title, form_id, version])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def write_xlsform(path: Path, **kwargs: Any) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(build_xlsform_bytes(**kwargs))
    return path


def instance_xml(
    *,
    form_id: str,
    version: str,
    instance_uuid: str,
    answers: dict[str, str],
) -> bytes:
    parts = [
        "<?xml version='1.0' encoding='UTF-8'?>",
        f'<data id="{form_id}" version="{version}">',
    ]
    for key, value in answers.items():
        if not key or value is None or value == "":
            continue
        safe = str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        parts.append(f"  <{key}>{safe}</{key}>")
    parts.append("  <meta>")
    parts.append(f"    <instanceID>uuid:{instance_uuid}</instanceID>")
    parts.append("  </meta>")
    parts.append("</data>")
    return "\n".join(parts).encode("utf-8")
