import io
import json
import logging
import uuid

from typing import Optional

import anthropic
import openpyxl

from django.conf import settings
from pydantic import BaseModel
from pyxform import create_survey_from_xls


logger = logging.getLogger(__name__)

XLSFORM_SYSTEM_PROMPT = """You are an expert ODK XLSForm designer. You create well-structured data collection forms based on user descriptions.

When the user asks you to create or modify a form, you MUST respond with valid JSON matching the schema below. Do NOT include any text outside the JSON block.

## Response JSON Schema
{
  "survey": [
    {
      "type": "text|integer|decimal|range|select_one <list>|select_multiple <list>|note|date|time|dateTime|geopoint|geotrace|geoshape|image|audio|video|file|barcode|calculate|acknowledge|hidden|begin group|end group|begin repeat|end repeat|start|end|today|deviceid|phonenumber|username|email",
      "name": "<unique_variable_name>",
      "label": "<display text>",
      "hint": "<optional hint>",
      "required": "yes or empty",
      "relevant": "<optional skip logic, e.g. ${field} = 'value'>",
      "constraint": "<optional validation, e.g. . > 0 and . < 150>",
      "constraint_message": "<optional error message>",
      "appearance": "<optional: minimal, multiline, likert, field-list, columns, etc>",
      "calculation": "<optional formula>",
      "default": "<optional default value>",
      "choice_filter": "<optional filter expression for cascading selects>",
      "parameters": "<optional: e.g. randomize=true, start=0 end=10 step=1, max-pixels=1000>",
      "repeat_count": "<optional: fixed or dynamic repeat count>",
      "read_only": "<optional: yes to make non-editable>",
      "media::image": "<optional: image filename>",
      "label::Language (code)": "<optional: translated label, e.g. label::French (fr)>"
    }
  ],
  "choices": [
    {
      "list_name": "<matches select type>",
      "name": "<stored value>",
      "label": "<display text>",
      "label::Language (code)": "<optional: translated label>"
    }
  ],
  "settings": {
    "form_title": "<human readable title>",
    "form_id": "<unique_snake_case_id>",
    "version": "1"
  },
  "message": "<Your explanation to the user about what you created/changed>"
}

## Rules
- Every `select_one <list>` or `select_multiple <list>` MUST have corresponding entries in `choices` with matching `list_name`.
- Use `begin group`/`end group` to organize related questions.
- Use `begin repeat`/`end repeat` for repeatable sections.
- Variable `name` fields must be unique, lowercase, use underscores, and start with a letter.
- Unless the user specifically asks, don't use the meta data fields: start, end, today, deviceid.
- Add `note` type questions for instructions or section headers.
- Use `relevant` for skip logic with `${variable_name}` references.
- Use `constraint` for data validation.
- When modifying an existing form, return the COMPLETE updated form, not just the changes.
- Preserve ALL existing columns/fields from the original form, including any you don't recognize.
- For internationalized forms, preserve all `label::Language` and `hint::Language` columns.
- Your `message` should briefly explain what you created or changed.
- Don't include by default a question to ask for health facility, region or district (org units in general) because the tool handling these forms does not need this.
- Make sure that no question has the same name as a previous question.
## Common patterns
- QR/barcode scanning: type "barcode"
- GPS location: type "geopoint"
- Photo capture: type "image"
- Repeating sections (e.g., multiple children): use "begin repeat"/"end repeat"
- Cascading selects: use choice_filter
- Read-only calculated fields: set read_only to "yes" with a calculation
"""


class SurveyRow(BaseModel, extra="allow"):
    type: str
    name: str
    label: str = ""


class ChoiceRow(BaseModel, extra="allow"):
    list_name: str
    name: str
    label: str = ""


class FormSettings(BaseModel, extra="allow"):
    form_title: str = ""
    form_id: str = ""
    version: str = "1"


class GeneratedForm(BaseModel):
    survey: list[SurveyRow]
    choices: list[ChoiceRow]
    settings: FormSettings
    message: str


def _parse_sheet(ws) -> list[dict]:
    """Parse a worksheet into a list of dicts, preserving ALL columns."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Preserve original header casing for the output
    raw_headers = rows[0]
    headers = []
    for h in raw_headers:
        headers.append(str(h).strip() if h else "")

    result = []
    for row in rows[1:]:
        row_dict = {}
        all_empty = True
        for i, header in enumerate(headers):
            if not header:
                continue
            val = row[i] if i < len(row) else None
            if val is not None:
                row_dict[header] = str(val)
                all_empty = False
            else:
                row_dict[header] = ""
        if not all_empty:
            result.append(row_dict)
    return result


def parse_xlsform_to_json(xls_file) -> dict:
    """Parse an existing XLSForm Excel file, preserving ALL columns.

    This allows loading an existing form into the conversation context
    so the AI can modify it without losing any data.
    """
    wb = openpyxl.load_workbook(xls_file, read_only=True)

    survey_rows = _parse_sheet(wb["survey"]) if "survey" in wb.sheetnames else []
    choice_rows = _parse_sheet(wb["choices"]) if "choices" in wb.sheetnames else []

    # Settings is a single-row sheet, parse it as a dict
    settings_data = {"form_title": "", "form_id": "", "version": "1"}
    if "settings" in wb.sheetnames:
        settings_rows = _parse_sheet(wb["settings"])
        if settings_rows:
            settings_data = settings_rows[0]

    wb.close()

    return {
        "survey": survey_rows,
        "choices": choice_rows,
        "settings": settings_data,
    }


def call_claude(message: str, conversation_history: list[dict], api_key: Optional[str] = None) -> dict:
    """Call Claude API with the conversation and return the raw response."""
    client = anthropic.Anthropic(api_key=api_key)

    messages = []
    for entry in conversation_history:
        messages.append({"role": entry["role"], "content": entry["content"]})
    messages.append({"role": "user", "content": message})

    response = client.messages.create(
        model=settings.FORM_COPILOT_MODEL,
        max_tokens=8192,
        system=XLSFORM_SYSTEM_PROMPT,
        messages=messages,
    )

    return response.content[0].text


def parse_form_response(response_text: str) -> GeneratedForm:
    """Parse Claude's response into a GeneratedForm."""
    text = response_text.strip()

    # Handle markdown code blocks
    if "```json" in text:
        text = text.split("```json")[1].split("```")[0].strip()
    elif "```" in text:
        text = text.split("```")[1].split("```")[0].strip()

    data = json.loads(text)
    return GeneratedForm(**data)


def _collect_all_keys(rows: list) -> list[str]:
    """Collect all unique keys from a list of dicts/models, preserving order."""
    seen = set()
    keys = []
    for row in rows:
        row_dict = row.model_dump() if hasattr(row, "model_dump") else row
        for key in row_dict:
            if key not in seen:
                seen.add(key)
                keys.append(key)
    return keys


def build_xlsform(form: GeneratedForm) -> io.BytesIO:
    """Build an XLSForm Excel file from the generated form data."""
    wb = openpyxl.Workbook()

    # Survey sheet
    survey_ws = wb.active
    survey_ws.title = "survey"
    survey_keys = _collect_all_keys(form.survey)
    survey_ws.append(survey_keys)
    for row in form.survey:
        row_dict = row.model_dump()
        survey_ws.append([row_dict.get(k, "") for k in survey_keys])

    # Choices sheet
    choices_ws = wb.create_sheet("choices")
    choices_keys = _collect_all_keys(form.choices)
    choices_ws.append(choices_keys)
    for row in form.choices:
        row_dict = row.model_dump()
        choices_ws.append([row_dict.get(k, "") for k in choices_keys])

    # Settings sheet
    settings_ws = wb.create_sheet("settings")
    settings_dict = form.settings.model_dump()
    settings_ws.append(list(settings_dict.keys()))
    settings_ws.append(list(settings_dict.values()))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def patch_xlsform_form_id(xlsform_buffer: io.BytesIO, form_odk_id: str) -> io.BytesIO:
    """Return a new BytesIO with the settings sheet form_id replaced."""
    xlsform_buffer.seek(0)
    wb = openpyxl.load_workbook(xlsform_buffer)
    if "settings" in wb.sheetnames:
        ws = wb["settings"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "form_id" in headers:
            col_index = headers.index("form_id") + 1  # 1-based
            # Update row 2 (first data row)
            ws.cell(row=2, column=col_index, value=form_odk_id)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def convert_to_xform_xml(xlsform_buffer: io.BytesIO) -> Optional[str]:
    """Convert an XLSForm to XForm XML using pyxform."""
    try:
        xlsform_buffer.seek(0)
        xlsform_buffer.name = "form.xlsx"
        survey = create_survey_from_xls(xlsform_buffer, default_name="data")
        return survey.to_xml(validate=False)
    except Exception as e:
        logger.error("Failed to convert XLSForm to XForm XML: %s", e)
        return None


def generate_form(
    message: str,
    conversation_history: list[dict],
    existing_form_odk_id: Optional[str] = None,
    api_key: Optional[str] = None,
) -> dict:
    """Call the AI and return the parsed form object plus updated conversation history.

    Returns a dict with:
    - assistant_message: The agent's response text
    - form: GeneratedForm instance (None if the response was conversational)
    - conversation_history: Updated conversation history
    """
    response_text = call_claude(message, conversation_history, api_key=api_key)

    new_history = list(conversation_history) + [
        {"role": "user", "content": message},
        {"role": "assistant", "content": response_text},
    ]

    try:
        form = parse_form_response(response_text)

        if existing_form_odk_id:
            # When editing an existing form, always use its original ODK form_id
            # so the /api/formversions/ consistency check passes.
            form.settings.form_id = existing_form_odk_id
        else:
            # No existing form context (new form, or form with no ODK id yet).
            # Generate a guaranteed-unique id so the uniqueness check never fails.
            form.settings.form_id = f"form_{uuid.uuid4().hex[:12]}"

        # Always clear the version so parse_xls_form auto-generates the correct
        # next version number. The AI always writes "1", which would fail the
        # "must be greater than previous version" check on any subsequent save.
        form.settings.version = ""

        return {
            "assistant_message": form.message,
            "form": form,
            "conversation_history": new_history,
        }

    except (json.JSONDecodeError, Exception) as e:
        logger.info("Response was not a form (might be conversational): %s", e)
        return {
            "assistant_message": response_text,
            "form": None,
            "conversation_history": new_history,
        }
