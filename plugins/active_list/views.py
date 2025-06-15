import calendar
import datetime
import hashlib
import math
import os
import tempfile
import uuid

from datetime import datetime, timedelta

import openpyxl
import pandas as pd
import pytz

from django.contrib.auth.decorators import login_required
from django.core.exceptions import FieldDoesNotExist
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import models
from django.db.models import Max
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.encoding import smart_str
from django_xlsform_validator.validation import XLSFormValidator

from iaso.models import Form, FormVersion, OrgUnit

from ..polio.settings import DISTRICT
from .forms import ActivePatientsListForm, ValidationForm
from .models import (
    HIV_HIV1,
    HIV_HIV1_AND_2,
    HIV_HIV2,
    HIV_UNKNOWN,
    INACTIVE_REASONS_DICT,
    PATIENT_HISTORY_DISPLAY_FIELDS,
    PATIENT_LIST_DISPLAY_FIELDS,
    SOURCE_EXCEL,
    TREATMENT_1STLINE,
    TREATMENT_2NDLINE,
    TREATMENT_3RDLINE,
    TREATMENT_LINE_UNKNOWN,
    VALIDATION_STATUS_CHOICES,
    Import,
    Patient,
    PatientInactiveEvent,
    Record,
    Validation,
)


UPLOAD_FOLDER = "upload"  # Create this folder in your project directory
FA_DISTRICT_ORG_UNIT_TYPE_ID = os.environ.get("FA_DISTRICT_ORG_UNIT_TYPE_ID", 4)
FA_HF_ORG_UNIT_TYPE_ID = os.environ.get("FA_HF_ORG_UNIT_TYPE_ID", 5)

ALLOWED_EXTENSIONS = {"xlsx", "xls", "ods", "csv", "xlsb", "xltm", "xltx", "xlsm"}
OK = "OK"
FILE_DATA_PROBLEM = "FILE_DATA_PROBLEM"
ERROR_FILE_ALREADY_UPLOADED = "ERROR_FILE_ALREADY_UPLOADED"
ERROR_PERIOD_ALREADY_UPLOADED = "ERROR_PERIOD_ALREADY_UPLOADED"
ERROR_NO_FILE = "ERROR_NO_FILE"
ERROR_WRONG_FILE_FORMAT = "ERROR_WRONG_FILE_FORMAT"
ERROR_WRONG_CODE = "ERROR_WRONG_CODE"
ERROR_WRONG_PERIOD = "ERROR_WRONG_PERIOD"
ERROR_UNKNOWN = "ERROR_UNKNOWN"

RESPONSES = {
    OK: {"message": "Merci! Votre fichier est bien enregistré!", "status": 201},
    ERROR_FILE_ALREADY_UPLOADED: {
        "message": "Ce fichier a déjà été uploadé, êtes vous sûr?",
        "status": 400,
        "bypassable": True,
    },
    ERROR_PERIOD_ALREADY_UPLOADED: {
        "message": "Cette période a déjà été encodée pour cet établissement, êtes vous sûr de vouloir continuer?",
        "status": 400,
        "bypassable": True,
    },
    ERROR_NO_FILE: {"message": "Vous n'avez pas ajouté de fichier", "status": 400, "bypassable": False},
    ERROR_WRONG_FILE_FORMAT: {"message": "Ce fichier n'est pas un fichier excel", "status": 400, "bypassable": True},
    ERROR_WRONG_CODE: {
        "message": "Le CODE ETS dans le fichier ne correspond pas à l'établissement sélectionné. Veuillez vérifier que vous avez sélectionné le bon établissement ou que le CODE ETS dans le fichier est correct.",
        "status": 400,
        "bypassable": False,
    },
    ERROR_WRONG_PERIOD: {
        "message": "La période dans le fichier ne correspond pas à la période sélectionnée. Veuillez vérifier que vous avez sélectionné la bonne période ou que la période dans le fichier est correcte.",
        "status": 400,
        "bypassable": False,
    },
    ERROR_UNKNOWN: {"message": "Erreur", "status": 400, "bypassable": False},
    FILE_DATA_PROBLEM: {
        "message": "Erreur dans les données du fichier, regardez la version annotée du fichier pour trouver les problèmes",
        "status": 400,
        "bypassable": False,
    },
}


@login_required
def homepage(request):
    return render(request, "homepage.html", {})


@login_required
def patient_list(request):
    return render(
        request,
        "patient_list.html",
        {
            "FA_HF_ORG_UNIT_TYPE_ID": FA_HF_ORG_UNIT_TYPE_ID,
            "FA_DISTRICT_ORG_UNIT_TYPE_ID": FA_DISTRICT_ORG_UNIT_TYPE_ID,
        },
    )


@login_required
def completeness(request, region_id, month):
    # Get the OrgUnit object, or raise a 404 error if it doesn't exist
    region = get_object_or_404(OrgUnit, id=region_id)
    districts = OrgUnit.objects.filter(parent=region).order_by("name")

    active_patients_list = Record.objects.filter(org_unit=org_unit, period=month).order_by("id")

    context = {
        "org_unit": region,
        "active_patients": active_patients_list,
        "PATIENT_LIST_DISPLAY_FIELDS": PATIENT_LIST_DISPLAY_FIELDS,
    }

    # Render the template with the context data
    return render(request, "completeness.html", context)


def convert_yyyymm_to_file_format(yyyymm_period):
    """Convert YYYY-MM format to Mon-YY format (e.g., 2024-06 -> Jun-24)"""
    try:
        from datetime import datetime

        # Parse the YYYY-MM format
        dt = datetime.strptime(yyyymm_period, "%Y-%m")
        # Format as Mon-YY (e.g., Jun-24)
        return dt.strftime("%b-%y")
    except ValueError:
        return yyyymm_period  # Return original if parsing fails


def convert_file_format_to_yyyymm(file_format_period):
    """Convert Mon-YY format to YYYY-MM format  (e.g.,Jun-24  -> 2024-06)"""
    try:
        from datetime import datetime

        # Parse the YYYY-MM format
        dt = datetime.strptime(file_format_period, "%b-%y")
        # Format as Mon-YY (e.g., Jun-24)
        return dt.strftime("%Y-%m")
    except ValueError:
        return file_format_period  # Return original if parsing fails


@login_required
def patient_history(request):
    identifier = request.GET.get("identifier", "")
    record_mode = request.GET.get("record_mode", "latest")
    if identifier:
        patient = get_object_or_404(Patient, identifier_code=identifier)
        if record_mode == "latest":
            latest_ids = Record.objects.filter(patient=patient).values("period").annotate(latest_id=Max("id"))
            records = Record.objects.filter(id__in=[item["latest_id"] for item in latest_ids]).order_by("-period")
        elif record_mode == "all":
            records = Record.objects.filter(patient=patient).order_by("-id")

        data = []
        for record in records:
            patient_object = {"Période": record.period[:7], "FOSA": record.org_unit.name}
            for field in PATIENT_HISTORY_DISPLAY_FIELDS:
                patient_object[PATIENT_HISTORY_DISPLAY_FIELDS[field]] = get_human_readable_value(record, field)
            is_last_record = record.patient.last_record.id == record.id

            patient_object["Dernier Enreg"] = "*" if is_last_record else ""
            patient_object["Date D'import"] = record.import_source.creation_date.strftime("%Y-%m-%d %H:%M ")
            patient_object["Importeur"] = record.import_source.user.username if record.import_source.user else "Inconnu"
            loss_text = ""
            first = True
            for event in record.loss_events.all():
                if not first:
                    loss_text += "\n "
                loss_text += "Date: %s - Raison: %s" % (
                    event.date.strftime("%Y-%m-%d"),
                    INACTIVE_REASONS_DICT[event.reason],
                )
                first = False
            patient_object["Perdu"] = loss_text
            data.append(patient_object)
        return render(request, "patient_history.html", {"data": data, "patient": patient, "record_mode": record_mode})

    return render(request, "patient_history.html", {"data": []})


@login_required
def patient_search(request):
    return render(request, "patient_search.html", {})


@login_required
def import_detail_view(request, import_id):
    """
    View to display a single Import object and its associated ActivePatientsList entries in a table.
    Args:
        request: The HTTP request object.
        import_id: The ID of the Import object to retrieve.
    Returns:
        A rendered HTML template with the Import object and its ActivePatientsList entries.
    """
    # Get the Import object, or raise a 404 error if it doesn't exist
    import_obj = get_object_or_404(Import, id=import_id)

    # Get the ActivePatientsList entries associated with the Import, ordered by some field (e.g., 'id')
    active_patients_list = Record.objects.filter(import_source=import_obj).order_by("id")

    context = {
        "import_obj": import_obj,
        "active_patients": active_patients_list,
        "PATIENT_LIST_DISPLAY_FIELDS": PATIENT_LIST_DISPLAY_FIELDS,
    }

    # Render the template with the context data
    return render(request, "import.html", context)


@login_required
def upload(request):
    if request.method == "POST":
        org_unit_id = request.POST.get("org_unit_id")
        period = request.POST.get("period")
        bypass = request.POST.get("bypass")

        if "file" not in request.FILES:
            result = ERROR_NO_FILE
        else:
            file = request.FILES["file"]
            if file.name == "":
                result = ERROR_NO_FILE
            elif not allowed_file(file.name):
                result = ERROR_WRONG_FILE_FORMAT
            elif file and allowed_file(file.name):
                filename = smart_str(file.name)
                result, annotated_file = handle_upload(filename, file, org_unit_id, period, bypass, request.user)
            else:
                result = ERROR_UNKNOWN
        answer = RESPONSES[result]
        if result == FILE_DATA_PROBLEM:
            validation_id = str(uuid.uuid4())

            # Save to shared storage for multi-server compatibility
            file_path = f"validation_files/validation_{validation_id}.xlsx"
            annotated_file.seek(0)
            stored_path = default_storage.save(file_path, ContentFile(annotated_file.read()))

            request.session[f"validation_{validation_id}"] = {"file_path": stored_path}
            answer["download_id"] = validation_id
        print("answer", answer)
        return JsonResponse(answer, status=RESPONSES[result]["status"])

    return render(
        request,
        "upload.html",
        {
            "FA_HF_ORG_UNIT_TYPE_ID": FA_HF_ORG_UNIT_TYPE_ID,
            "FA_DISTRICT_ORG_UNIT_TYPE_ID": FA_DISTRICT_ORG_UNIT_TYPE_ID,
        },
    )


@login_required
def download(request):
    validation_id = request.GET.get("id")
    if not validation_id:
        raise Http404("Download ID not provided")

    session_key = f"validation_{validation_id}"
    validation_data = request.session.get(session_key)

    if not validation_data or not default_storage.exists(validation_data["file_path"]):
        raise Http404("File not found or expired")

    file_obj = default_storage.open(validation_data["file_path"], "rb")
    response = FileResponse(
        file_obj,
        as_attachment=True,
        filename="highlighted_spreadsheet.xlsx",
    )
    return response


@login_required
def validation(request):
    org_unit = None
    if request.method == "POST":
        form = ValidationForm(request.POST)
        if form.is_valid():
            form.instance.user_id = request.user.id
            form.instance.user_name = request.user.username
            form.instance.level = DISTRICT
            form.instance.save()
            # return redirect('validation_list') # Replace 'validation_list' with the actual URL name
    else:
        form = ValidationForm()
    return render(
        request,
        "validation.html",
        {
            "org_unit": org_unit,
            "statuses": VALIDATION_STATUS_CHOICES,
            "request": request,
            "form": form,
            "FA_HF_ORG_UNIT_TYPE_ID": FA_HF_ORG_UNIT_TYPE_ID,
            "FA_DISTRICT_ORG_UNIT_TYPE_ID": FA_DISTRICT_ORG_UNIT_TYPE_ID,
        },
    )


def patient_form(request, patient_id=None):
    """
    View for creating or updating an ActivePatientsList instance.
    """
    if patient_id:
        patient = get_object_or_404(Record, pk=patient_id)
        form = ActivePatientsListForm(request.POST or None, instance=patient)
    else:
        form = ActivePatientsListForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("patient_list")  # Replace 'patient_list' with your list view name

    res = render(request, "patient_form.html", {"form": form, "patient_id": patient_id})

    return res


def active_count(org_unit_id):
    return Patient.objects.filter(active=True).filter(last_record__org_unit_id=org_unit_id).count()


def received_count(org_unit_id, month):
    """
    Returns the count of ActivePatientsList entries for a given org_unit and month.
    """
    latest_ids = Record.objects.filter(org_unit_id=org_unit_id).values("patient_id").annotate(latest_id=Max("id"))

    records = Record.objects.filter(id__in=[item["latest_id"] for item in latest_ids])
    records = filter_received_patients_for_the_month(records, month)
    return records.count()


def stats(org_unit_id, month):
    """
    Returns the count of ActivePatientsList entries for a given org_unit and month.
    """
    previous_month = get_previous_period(month) + "-01"
    anteprevious_month = get_previous_period(month) + "-01"
    month = month + "-01"
    latest_ids = Record.objects.filter(org_unit_id=org_unit_id).values("patient_id").annotate(latest_id=Max("id"))

    records = Record.objects.filter(id__in=[item["latest_id"] for item in latest_ids])

    """
    Perdus de Vue, Décédés et Arrêts de Traitement
    Nouvelles inclusions
    Patient de retour après interruption
    This is not correct, we need to check better this code
    """
    lost = (
        PatientInactiveEvent.objects.filter(org_unit_id=org_unit_id)
        .filter(date__gte=previous_month)
        .filter(date__lt=month)
        .count()
    )
    previous_lost = (
        PatientInactiveEvent.objects.filter(org_unit_id=org_unit_id)
        .filter(date__gte=anteprevious_month)
        .filter(date__lt=previous_month)
        .count()
    )

    records = filter_received_patients_for_the_month(records, month)
    previous_records = filter_received_patients_for_the_month(records, previous_month)
    deceased = records.filter(death=True).count()
    previous_deceased = previous_records.filter(death=True).count()
    stopped = records.filter(art_stoppage=True).count()
    previous_stopped = previous_records.filter(art_stoppage=True).count()
    new = records.filter(new_inclusion=True).count()
    previous_new = previous_records.filter(new_inclusion=True).count()
    return_to_care = records.filter(return_to_care=True).count()
    previous_return_to_care = previous_records.filter(return_to_care=True).count()

    return {
        "Perdu de vue": "%d (%d)" % (lost, previous_lost),
        "Décédé": "%d (%d)" % (deceased, previous_deceased),
        "Arrêt TAR": "%d (%d)" % (stopped, previous_stopped),
        "Nouveau": "%d (%d)" % (new, previous_new),
        "Retour en soin": "%d (%d)" % (return_to_care, previous_return_to_care),
    }


def get_previous_period(period):
    """
    Computes the 'yy-mm' format for the month preceding the given period.

    Args:
        period (str): The current period in "yy-mm" format.

    Returns:
        str: The previous month in "yy-mm" format.
    """

    # Convert the period string to a datetime object
    current_date = datetime.strptime(period, "%Y-%m")

    # Subtract one month using timedelta
    first_of_current_month = current_date.replace(day=1)
    last_of_previous_month = first_of_current_month - timedelta(days=1)

    # Format the result back to "yy-mm"
    previous_period = last_of_previous_month.strftime("%Y-%m")
    return previous_period


@login_required
def validation_api(request, org_unit_id, month):
    org_units = OrgUnit.objects.filter(parent_id=org_unit_id).order_by("name")
    previous_period = get_previous_period(month)
    table_content = []
    report_count = 0
    validation_count = 0
    for org_unit in org_units:
        obj = {}
        obj["Site"] = org_unit.name
        latest_import = Import.objects.filter(org_unit=org_unit).filter(month=month).order_by("-creation_date").first()

        obj["A rapporté"] = "Non"
        obj["Dernier rapport"] = ""
        obj["Actifs"] = active_count(org_unit.id)
        obj["Reçus"] = ""
        obj.update(stats(org_unit.id, month))
        obj["Date Validation"] = ""
        obj["Statut"] = ""
        obj["Observation"] = ""
        obj["Validateur"] = ""

        if latest_import:
            report_count = report_count + 1
            obj["A rapporté"] = "Oui"
            obj["Dernier rapport"] = latest_import.creation_date.strftime("%d/%m/%y %H:%M")

        latest_validation = (
            Validation.objects.filter(period=month).filter(org_unit_id=org_unit.id).order_by("-created_at").first()
        )
        if latest_validation:
            validation_count = validation_count + 1

            obj["Reçus"] = "%d (%d)" % (
                received_count(org_unit.id, month),
                received_count(org_unit.id, previous_period),
            )
            obj["Date Validation"] = latest_validation.created_at.strftime("%d/%m/%y %H:%M")
            obj["Statut"] = "Validé" if latest_validation.validation_status == "OK" else "Invalide"
            obj["Observation"] = latest_validation.comment
            obj["Validateur"] = latest_validation.user_name

        else:
            obj["Statut"] = ""
            obj["Observation"] = ""
            obj["Validateur"] = ""

        obj["org_unit_id"] = org_unit.id

        table_content.append(obj)

    res = {
        "table_content": table_content,
        "completeness": "Rapports: %d/%d - Validations: %d/%d"
        % (report_count, len(org_units), validation_count, len(org_units)),
    }
    return JsonResponse(res, status=200, safe=False)


def get_human_readable_value(model_instance, field_name):
    """
    Returns the human-readable value of a choice field in a Django model.

    Args:
      model_instance: An instance of a Django model.
      field_name: The name of the choice field.

    Returns:
      The human-readable value of the field, or None if the field is not a choice field
      or does not exist.
    """
    try:
        field = model_instance._meta.get_field(field_name)
        if field.choices:
            method_name = f"get_{field_name}_display"
            return getattr(model_instance, method_name)()
        if isinstance(field, models.BooleanField):
            return "Oui" if getattr(model_instance, field_name) else "Non"
        print(field, field_name)
        if getattr(model_instance, field_name) is None:
            return ""
    except FieldDoesNotExist:
        return None

    return getattr(model_instance, field_name)


def get_first_and_last_day(date_str):
    """
    Calculates the first and last day of the month given a date string in the format "YYYY-MM".

    Args:
      date_str: A string representing the year and month in the format "YYYY-MM".

    Returns:
      A tuple containing two strings:
        - The first day of the month in the format "YYYY-MM-DD".
        - The last day of the month in the format "YYYY-MM-DD".
    """
    year, month = map(int, date_str[:7].split("-"))

    # Get the number of days in the month
    _, last_day = calendar.monthrange(year, month)

    # Format the first and last days
    first_day_str = f"{year}-{month:02d}-01"
    last_day_str = f"{year}-{month:02d}-{last_day}"

    return first_day_str, last_day_str


def filter_expected_patients_for_the_month(patients, month):
    first_day_str, last_day_str = get_first_and_last_day(month)
    return patients.filter(
        next_dispensation_date__gte=first_day_str
    ).filter(
        next_dispensation_date__lte=last_day_str
    )  # todo improve here to only show the most recent record for each patient (maybe also filter out inactive patients?)


def filter_received_patients_for_the_month(patients, month):
    first_day_str, last_day_str = get_first_and_last_day(month)
    return patients.filter(
        last_dispensation_date__gte=first_day_str
    ).filter(
        last_dispensation_date__lte=last_day_str
    )  # todo improve here to only show the most recent record for each patient (maybe also filter out inactive patients?)


@login_required
def patient_list_api(request, org_unit_id, month):
    format = request.GET.get("format", "json")
    mode = request.GET.get("mode", "default")
    last_import = None
    if mode == "default":
        last_import = Import.objects.filter(org_unit=org_unit_id)
        last_import = last_import.filter(month=month).order_by("-creation_date").first()
        records = Record.objects.filter(import_source=last_import).order_by("number")
    elif mode == "expected":
        records = Record.objects.filter(org_unit_id=org_unit_id)
        records = filter_expected_patients_for_the_month(records, month).order_by("next_dispensation_date")
    elif mode == "active":
        record_ids = (
            Patient.objects.filter(active=True)
            .filter(last_record__org_unit_id=org_unit_id)
            .values_list("last_record__id", flat=True)
        )
        records = Record.objects.filter(id__in=record_ids).order_by("number")
    elif mode == "lost":
        first_day, last_day = get_first_and_last_day(month)
        inactive_events = (
            PatientInactiveEvent.objects.filter(org_unit_id=org_unit_id)
            .filter(date__gte=first_day)
            .filter(date__lte=last_day)
        )
        records = [event.last_patient_record_at_time_of_loss for event in inactive_events]

    if format == "xls":
        org_unit = OrgUnit.objects.get(pk=org_unit_id)
        return export_active_patients_excel(records, org_unit.name, month)
    table_content = []
    for record in records:
        patient_object = {
            "Code identifiant": record.patient.identifier_code,
            "Actif": "Oui" if record.patient.active else "Non",
        }

        for field in PATIENT_LIST_DISPLAY_FIELDS:
            patient_object[PATIENT_LIST_DISPLAY_FIELDS[field]] = get_human_readable_value(record, field)
            if mode in ("expected", "lost", "active"):
                if field == "days_dispensed":
                    patient_object["Date de rupture"] = get_human_readable_value(record, "next_dispensation_date")
        patient_object["Voir"] = (
            '<a href="/active_list/patient_history/?identifier=%s">Voir</a>' % record.patient.identifier_code
        )
        table_content.append(patient_object)
    res = {
        "table_content": table_content,
    }
    if format == "json":
        return JsonResponse(res, status=200, safe=False)
    return render(request, "partials/patient_table.html", {"data": table_content, "last_import": last_import})


@login_required
def patient_list_upload_format_api(request, org_unit_id, month):
    """API endpoint to export patient data in upload-compatible Excel format"""
    mode = request.GET.get("mode", "default")
    last_import = None

    # todo :  Same data retrieval logic as patient_list_api.Should not be replicated
    if mode == "default":
        last_import = Import.objects.filter(org_unit=org_unit_id)
        last_import = last_import.filter(month=month).order_by("-creation_date").first()
        records = Record.objects.filter(import_source=last_import).order_by("number")
    elif mode == "expected":
        records = Record.objects.filter(org_unit_id=org_unit_id)
        records = filter_expected_patients_for_the_month(records, month).order_by("next_dispensation_date")
    elif mode == "active":
        record_ids = (
            Patient.objects.filter(active=True)
            .filter(last_record__org_unit_id=org_unit_id)
            .values_list("last_record__id", flat=True)
        )
        records = Record.objects.filter(id__in=record_ids).order_by("number")
    elif mode == "lost":
        first_day, last_day = get_first_and_last_day(month)
        inactive_events = (
            PatientInactiveEvent.objects.filter(org_unit_id=org_unit_id)
            .filter(date__gte=first_day)
            .filter(date__lte=last_day)
        )
        records = [event.last_patient_record_at_time_of_loss for event in inactive_events]

    # Always export in upload-compatible Excel format
    org_unit = OrgUnit.objects.get(pk=org_unit_id)
    return export_active_patients_upload_format(records, org_unit.name, month)


@login_required
def patient_search_api(request):
    """
    View to handle search functionality for patients.
    """
    # Perform search logic here (e.g., filter ActivePatientsList based on query)
    query = request.GET.get("query", "")
    patients = (
        Patient.objects.filter(identifier_code__icontains=query)
        .distinct("identifier_code")
        .order_by("identifier_code")[:20]
    )
    table_content = []
    for patient in patients:
        patient_object = {}
        last_record = patient.last_record

        patient_object["Code identificateur"] = patient.identifier_code
        patient_object["Sexe"] = get_human_readable_value(last_record, "sex")
        patient_object["Âge"] = get_human_readable_value(last_record, "age")
        patient_object["Poids"] = get_human_readable_value(last_record, "weight")

        patient_object["Voir"] = (
            '<a href="/active_list/patient_history/?identifier=%s">Voir</a>' % patient.identifier_code
        )
        table_content.append(patient_object)
    # Render the results in a template
    return render(request, "partials/patient_table.html", {"data": table_content})


def export_active_patients_excel(active_patients, name, period):
    """Export patients in current display format (original functionality)"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add headers using French names from PATIENT_LIST_DISPLAY_FIELDS
    headers = list(PATIENT_LIST_DISPLAY_FIELDS.values())
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num).value = header

    for row_num, patient in enumerate(active_patients, 2):
        for col_num, field in enumerate(PATIENT_LIST_DISPLAY_FIELDS.keys(), 1):
            worksheet.cell(row=row_num, column=col_num).value = get_human_readable_value(patient, field)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="active_patients-%s-%s.xlsx"' % (name, period)
    workbook.save(response)
    return response


def export_active_patients_upload_format(active_patients, name, period):
    """Export patients in upload-compatible format with styling"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Define upload-compatible headers (matching the upload format)
    upload_headers = [
        "N°",
        "CODE ETS",
        "SITES",
        "Periode",
        "REGION",
        "DISTRICT",
        "CODE IDENTIFIANT",
        "SEXE",
        "AGE",
        "POIDS",
        "Nouvelle inclusion",
        "Transfert-In",
        "Retour dans les soins",
        "TB / VIH",
        "Type de VIH",
        "Ligne thérapeutique",
        "Date de la dernière dispensation",
        "Nombre de jours dispensés",
        "REGIME",
        "STABLE",
        "Transfert Out",
        "Décès",
        "Arrêt TARV",
        "Servi ailleurs",
    ]

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Light border style
    thin_border = Border(
        left=Side(style="thin", color="D1D1D1"),
        right=Side(style="thin", color="D1D1D1"),
        top=Side(style="thin", color="D1D1D1"),
        bottom=Side(style="thin", color="D1D1D1"),
    )

    # Data alignment styles
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    # Add headers to worksheet with styling
    for col_num, header in enumerate(upload_headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Helper function to convert YYYY-MM period to Mon-YY format
    def convert_yyyymm_to_file_format(yyyymm_period):
        """Convert YYYY-MM format to Mon-YY format (e.g., 2024-06 -> Jun-24)"""
        try:
            from datetime import datetime

            # Parse the YYYY-MM format
            dt = datetime.strptime(yyyymm_period, "%Y-%m")
            # Format as Mon-YY (e.g., Jun-24)
            return dt.strftime("%b-%y")
        except ValueError:
            return yyyymm_period  # Return original if parsing fails

    # Convert period to upload format
    file_period = convert_yyyymm_to_file_format(period)

    # Helper function to style data cells
    def style_data_cell(cell, use_center=False):
        cell.border = thin_border
        cell.alignment = center_alignment if use_center else left_alignment

    # Add data rows with styling
    for row_num, patient in enumerate(active_patients, 2):
        # N° (centered)
        cell = worksheet.cell(row=row_num, column=1)
        cell.value = getattr(patient, "number", row_num - 1)
        style_data_cell(cell, use_center=True)

        # CODE ETS (centered)
        cell = worksheet.cell(row=row_num, column=2)
        cell.value = getattr(patient, "code_ets", "")
        style_data_cell(cell, use_center=True)

        # SITES (left aligned) - use org_unit.name
        cell = worksheet.cell(row=row_num, column=3)
        try:
            sites_name = patient.org_unit.name if patient.org_unit else ""
        except AttributeError:
            sites_name = getattr(patient, "facility_name", "")
        cell.value = sites_name
        style_data_cell(cell)

        # Periode (centered)
        cell = worksheet.cell(row=row_num, column=4)
        cell.value = file_period
        style_data_cell(cell, use_center=True)

        # REGION (left aligned) - use org_unit.parent.parent.name
        cell = worksheet.cell(row=row_num, column=5)
        try:
            region_name = ""
            if patient.org_unit and patient.org_unit.parent and patient.org_unit.parent.parent:
                region_name = patient.org_unit.parent.parent.name
        except AttributeError:
            region_name = getattr(patient, "region", "")
        cell.value = region_name
        style_data_cell(cell)

        # DISTRICT (left aligned) - use org_unit.parent.name
        cell = worksheet.cell(row=row_num, column=6)
        try:
            district_name = ""
            if patient.org_unit and patient.org_unit.parent:
                district_name = patient.org_unit.parent.name
        except AttributeError:
            district_name = getattr(patient, "district", "")
        cell.value = district_name
        style_data_cell(cell)

        # CODE IDENTIFIANT (centered)
        cell = worksheet.cell(row=row_num, column=7)
        cell.value = getattr(patient.patient, "identifier_code", "") if hasattr(patient, "patient") else ""
        style_data_cell(cell, use_center=True)

        # SEXE (centered)
        cell = worksheet.cell(row=row_num, column=8)
        cell.value = get_human_readable_value(patient, "sex")
        style_data_cell(cell, use_center=True)

        # AGE (centered)
        cell = worksheet.cell(row=row_num, column=9)
        cell.value = get_human_readable_value(patient, "age")
        style_data_cell(cell, use_center=True)

        # POIDS (centered)
        cell = worksheet.cell(row=row_num, column=10)
        cell.value = get_human_readable_value(patient, "weight")
        style_data_cell(cell, use_center=True)

        # Nouvelle inclusion (centered)
        cell = worksheet.cell(row=row_num, column=11)
        cell.value = 1 if getattr(patient, "new_inclusion", False) else 0
        style_data_cell(cell, use_center=True)

        # Transfert-In (centered)
        cell = worksheet.cell(row=row_num, column=12)
        cell.value = 1 if getattr(patient, "transfer_in", False) else 0
        style_data_cell(cell, use_center=True)

        # Retour dans les soins (centered)
        cell = worksheet.cell(row=row_num, column=13)
        cell.value = 1 if getattr(patient, "return_to_care", False) else 0
        style_data_cell(cell, use_center=True)

        # TB / VIH (centered)
        cell = worksheet.cell(row=row_num, column=14)
        cell.value = 1 if getattr(patient, "tb_hiv", False) else 0
        style_data_cell(cell, use_center=True)

        # Type de VIH (centered)
        cell = worksheet.cell(row=row_num, column=15)
        cell.value = get_human_readable_value(patient, "hiv_type")
        style_data_cell(cell, use_center=True)

        # Ligne thérapeutique (centered)
        cell = worksheet.cell(row=row_num, column=16)
        cell.value = get_human_readable_value(patient, "treatment_line")
        style_data_cell(cell, use_center=True)

        # Date de la dernière dispensation (centered)
        cell = worksheet.cell(row=row_num, column=17)
        cell.value = get_human_readable_value(patient, "last_dispensation_date")
        style_data_cell(cell, use_center=True)

        # Nombre de jours dispensés (centered)
        cell = worksheet.cell(row=row_num, column=18)
        cell.value = get_human_readable_value(patient, "days_dispensed")
        style_data_cell(cell, use_center=True)

        # REGIME (left aligned)
        cell = worksheet.cell(row=row_num, column=19)
        cell.value = getattr(patient, "regimen", "")
        style_data_cell(cell)

        # STABLE (centered)
        cell = worksheet.cell(row=row_num, column=20)
        stable_val = getattr(patient, "stable", False)
        cell.value = "Oui" if stable_val else "Non"
        style_data_cell(cell, use_center=True)

        # Transfert Out (centered)
        cell = worksheet.cell(row=row_num, column=21)
        cell.value = 1 if getattr(patient, "transfer_out", False) else 0
        style_data_cell(cell, use_center=True)

        # Décès (centered)
        cell = worksheet.cell(row=row_num, column=22)
        cell.value = 1 if getattr(patient, "death", False) else 0
        style_data_cell(cell, use_center=True)

        # Arrêt TARV (centered)
        cell = worksheet.cell(row=row_num, column=23)
        cell.value = 1 if getattr(patient, "art_stoppage", False) else 0
        style_data_cell(cell, use_center=True)

        # Servi ailleurs (centered)
        cell = worksheet.cell(row=row_num, column=24)
        cell.value = 1 if getattr(patient, "served_elsewhere", False) else 0
        style_data_cell(cell, use_center=True)

    # Auto-adjust column widths for better readability
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Set a reasonable width (between 10 and 20 characters)
        adjusted_width = min(max(max_length + 2, 10), 20)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Freeze the header row for easier navigation
    worksheet.freeze_panes = "A2"

    # Set worksheet tab name
    worksheet.title = f"Patients {period}"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="upload_format-%s-%s.xlsx"' % (name, period)
    workbook.save(response)
    return response


def check_presence(file_hash, org_unit_id, month):
    hash_exists = Import.objects.filter(hash_key=file_hash).exists()
    period_exists = Import.objects.filter(org_unit_id=org_unit_id, month=month).exists()

    if hash_exists:
        return "ERROR_FILE_ALREADY_UPLOADED"

    if period_exists:
        return "ERROR_PERIOD_ALREADY_UPLOADED"
    return "OK"


def hash_blob(binary_blob):
    """
    Calculates the SHA-256 hash of a binary blob.

    Args:
      binary_blob: The binary data as bytes.

    Returns:
      The hexadecimal representation of the SHA-256 hash.
    """

    hasher = hashlib.sha256()
    hasher.update(binary_blob)
    return hasher.hexdigest()


def handle_upload(file_name, file, org_unit_id, month, bypass=False, user=None):
    validator = XLSFormValidator()
    form = Form.objects.get(form_id="file_active_excel_validation")

    form_version = FormVersion.objects.filter(form=form).order_by("created_at").last()

    # Create a copy of the form file content to avoid pyxform closing the original file
    form_version.xls_file.seek(0)
    form_file_content = form_version.xls_file.read()

    # Create a fresh file object for pyxform validation

    validation_file = SimpleUploadedFile(
        "validation_form.xlsx",
        form_file_content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # Use temporary directory for multi-server compatibility
    with tempfile.TemporaryDirectory() as temp_dir:
        forms_dir = os.path.join(temp_dir, "forms")
        os.makedirs(forms_dir, exist_ok=True)
        validator.parse_xlsform(validation_file)

    # Create a copy of the upload file content to avoid pyxform closing the original file
    file.seek(0)
    upload_file_content = file.read()

    # Create a fresh file object for validation
    validation_upload_file = SimpleUploadedFile(
        file_name,
        upload_file_content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    result = validator.validate_spreadsheet(validation_upload_file)
    print("validation result", result)
    if not result["is_valid"]:
        # Create another fresh file object for error highlighting
        error_file_obj = SimpleUploadedFile(
            file_name,
            upload_file_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        error_file = validator.create_highlighted_excel(error_file_obj, result["errors"])
        print(FILE_DATA_PROBLEM, error_file)
        return FILE_DATA_PROBLEM, error_file

    content = upload_file_content
    h = hash_blob(content)
    result = check_presence(h, org_unit_id, month)
    # TODO store all the errors of the file and not only one of the error results
    file_check = result
    name_exists = Import.objects.filter(file_name=file_name).exists()
    if name_exists:
        base, ext = os.path.splitext(file_name)  # Split into base name and extension
        file_name = f"{base}_{uuid.uuid4()}{ext}"
    if bypass:
        result = OK
    if result == OK:
        # Create a fresh file object for Import model to avoid closed file issues
        import_file = SimpleUploadedFile(
            file_name,
            upload_file_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        i = Import(
            hash_key=h,
            file_name=file_name,
            org_unit_id=org_unit_id,
            month=month,
            file_check=file_check,
            source=SOURCE_EXCEL,
            file=import_file,
            user=user,
        )
        i.save()
        # Pass the content as BytesIO for pandas to read
        from io import BytesIO

        validation_result = import_data(BytesIO(content), i)
        # Check if validation failed during import
        if isinstance(validation_result, dict):
            # Delete the import since validation failed
            i.delete()

            error_type = validation_result.get("error")

            if error_type == ERROR_WRONG_CODE:
                result = ERROR_WRONG_CODE
                # Create dynamic error message with available aliases
                invalid_codes = validation_result.get("invalid_codes", [])
                invalid_codes = [str(code) for code in invalid_codes if code]  # Filter out empty codes
                available_aliases = validation_result.get("available_aliases", [])
                org_unit_name = validation_result.get("org_unit_name", "")

                # Update the error message to include specific information
                aliases_text = ", ".join(available_aliases) if available_aliases else "aucun alias défini"
                codes_text = ", ".join(invalid_codes)

                RESPONSES[ERROR_WRONG_CODE]["message"] = (
                    f"Le(s) CODE ETS '{codes_text}' dans le fichier ne correspond(ent) pas à l'établissement "
                    f"sélectionné '{org_unit_name}'. Les codes autorisés sont: {aliases_text}. "
                    f"Veuillez vérifier que vous avez sélectionné le bon établissement ou que le CODE ETS "
                    f"dans le fichier est correct."
                )

            elif error_type == ERROR_WRONG_PERIOD:
                result = ERROR_WRONG_PERIOD
                # Create dynamic error message with period details
                file_periods = validation_result.get("file_periods", [])
                selected_period_yyyymm = validation_result.get("selected_period_yyyymm", "")
                expected_file_period = validation_result.get("expected_file_period", "")
                org_unit_name = validation_result.get("org_unit_name", "")

                # Update the error message to include specific information
                periods_text = ", ".join([str(p) for p in file_periods])

                RESPONSES[ERROR_WRONG_PERIOD]["message"] = (
                    f"La(les) période(s) '{periods_text}' dans le fichier ne correspond(ent) pas à la période "
                    f"sélectionnée '{expected_file_period}' (basée sur {selected_period_yyyymm}) pour l'établissement '{org_unit_name}'. "
                    f"Veuillez vérifier que vous avez sélectionné la bonne période ou que la période "
                    f"dans le fichier est correcte (format attendu: {expected_file_period})."
                )
    return result, None


def validate_import(the_import):
    month = the_import.month
    import_lines = Record.objects.filter(import_source=the_import)
    viewed_patients_count = filter_expected_patients_for_the_month(import_lines, month).count()
    latest_ids = (
        Record.objects.filter(org_unit_id=the_import.org_unit_id)
        .values("identifier_code")
        .annotate(latest_id=Max("id"))
    )

    active_list = Record.objects.filter(id__in=[item["latest_id"] for item in latest_ids])

    patients = patients.filter(active=True).filter(org_unit_id=org_unit_id)


def check_file(file):
    # To implement
    # warning on multiple sheets
    # needed columns are not there -> refuse file
    # unrecognized value in column
    # wrong value in column (wrong type, wrong choice in a selection
    # name of facility does not match with the one that was picked
    pass


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def month_string_to_utc_timestamp(month_string):
    """
    Converts a month string in "YYYY-MM" format to a UTC timestamp representing the first day of the month.
    """
    year, month = map(int, month_string.split("-"))
    dt = datetime.datetime(year, month, 1, tzinfo=pytz.utc)
    timestamp = calendar.timegm(dt.utctimetuple())
    return dt


def import_data(file, the_import):
    # the code of this function is relatively bad, it has been generated by AI
    pd.set_option("display.max_rows", None)
    pd.set_option("display.max_columns", None)

    # Read the CSV file into a DataFrame
    df = pd.read_excel(file, sheet_name=0)
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)  # trimming all white spaces

    # Get the org unit for alias validation
    org_unit = the_import.org_unit
    # Rename columns
    df = df.rename(
        columns={
            "N°": "number",
            "CODE ETS": "code_ets",
            "CODE": "code_ets",
            "NOM ETABLISSEMENT": "facility_name",
            "SITES": "facility_name",
            "Periode": "period",
            "MOIS DE RAPPORTAGE": "period",
            "CODE IDENTIFIANT": "identifier_code",
            "SEXE": "sex",
            "POIDS": "weight",
            "Nouvelle inclusion": "new_inclusion",
            "Transfert-In": "transfer_in",
            "Retour dans les soins": "return_to_care",
            "TB / VIH": "tb_hiv",
            "Type de VIH": "hiv_type",
            "Régime": "REGIME",
            "régime": "REGIME",
            "regime": "REGIME",
            "Ligne therapeuthique": "treatment_line",
            "Ligne thérapeutique": "treatment_line",
            "Date de la dernière dispensation": "last_dispensation_date",
            "Nombre de jours dispensés": "days_dispensed",
            "STABLE": "stable",
            "Transfert Out": "transfer_out",
            "Décès": "death",
            "Arrêt TARV": "art_stoppage",
            "Servi ailleurs": "served_elsewhere",
        }
    )

    # Convert columns to boolean
    for col in [
        "new_inclusion",
        "transfer_in",
        "return_to_care",
        "tb_hiv",
        "transfer_out",
        "death",
        "art_stoppage",
        "served_elsewhere",
    ]:
        df[col] = df[col].apply(lambda x: True if x == 1 else False)

    # Map values in column `stable`
    df["stable"] = df["stable"].replace({"Oui": True, "Non": False})
    df["sex"] = df["sex"].replace({"F": "FEMALE", "M": "MALE"})
    df["treatment_line"] = df["treatment_line"].replace(
        {
            "Ligne 1": TREATMENT_1STLINE,
            "Ligne 2": TREATMENT_2NDLINE,
            "Ligne 3": TREATMENT_3RDLINE,
            "1": TREATMENT_1STLINE,
            "2": TREATMENT_2NDLINE,
            "3": TREATMENT_3RDLINE,
            1: TREATMENT_1STLINE,
            2: TREATMENT_2NDLINE,
            3: TREATMENT_3RDLINE,
            math.nan: TREATMENT_LINE_UNKNOWN,
        }
    )

    df["REGIME"] = df["REGIME"].replace(
        {
            "TDF 3TC DTG": "TDF/3TC/DTG",
            "ABC 3TC DTG": "ABC/3TC/DTG",
            "ABC 3TC ATV-r": "ABC/3TC/ATV-r",
            "ABC 3TC EFV": "ABC/3TC/EFV",
            "ABC 3TC LPV-r": "ABC/3TC/LPV-r",
            "ABC 3TC NVP": "ABC/3TC/NVP",
            "AZT 3TC ABC": "AZT/3TC/ABC",
            "AZT 3TC ATV-r": "AZT/3TC/ATV-r",
            "AZT 3TC DTG": "AZT/3TC/DTG",
            "AZT 3TC EFV": "AZT/3TC/EFV",
            "AZT 3TC LPV-r": "AZT/3TC/LPV-r",
            "AZT 3TC TDF": "AZT/3TC/TDF",
            "TDF 3TC ATV-r": "TDF/3TC/ATV-r",
            "TDF 3TC EFV": "TDF/3TC/EFV",
            "TDF 3TC LPV-r": "TDF/3TC/LPV-r",
        }
    )

    df["hiv_type"] = df["hiv_type"].replace(
        {
            1: HIV_HIV1,
            2: HIV_HIV2,
            "1&2": HIV_HIV1_AND_2,
            math.nan: HIV_UNKNOWN,
            "nan": HIV_UNKNOWN,
            "VIH 1 + 2": HIV_HIV1_AND_2,
            "VIH 1": HIV_HIV1,
            "VIH 2": HIV_HIV2,
            0: HIV_UNKNOWN,
        }
    )

    # Convert `last_dispensation_date` to datetime
    df["last_dispensation_date"] = pd.to_datetime(df["last_dispensation_date"])

    # Create a list of dictionaries
    data = []
    for _, row in df.iterrows():
        data.append(
            {
                "number": row["number"],
                "region": row["REGION"],
                "district": row["DISTRICT"],
                "code_ets": row["code_ets"],
                "facility_name": row.get("facility_name") if row.get("facility_name") else "",
                "period": row["period"],
                "identifier_code": row["identifier_code"].lower(),
                "sex": row["sex"],
                "age": row["AGE"],
                "weight": row["weight"] if not math.isnan(row["weight"]) else None,
                "new_inclusion": row["new_inclusion"],
                "transfer_in": row["transfer_in"],
                "return_to_care": row["return_to_care"],
                "tb_hiv": row["tb_hiv"],
                "hiv_type": row["hiv_type"],
                "treatment_line": row["treatment_line"],
                "last_dispensation_date": row["last_dispensation_date"],
                "days_dispensed": row["days_dispensed"],
                "regimen": row["REGIME"],
                "stable": row["stable"],
                "transfer_out": row["transfer_out"],
                "death": row["death"],
                "art_stoppage": row["art_stoppage"],
                "served_elsewhere": row["served_elsewhere"],
            }
        )

    # Validate CODE ETS against org unit aliases
    if data:  # Only validate if there's data
        # Get all unique CODE ETS values from the data
        code_ets_values = {row["code_ets"] for row in data if row["code_ets"]}

        # Get org unit aliases (empty list if None)
        org_unit_aliases = org_unit.aliases or []
        print(org_unit_aliases)
        # Check if any CODE ETS doesn't match org unit aliases
        for code_ets in code_ets_values:
            if str(code_ets) not in org_unit_aliases:
                print(f"CODE ETS validation failed: '{code_ets}' not found in org unit aliases: {org_unit_aliases}")
                # Return detailed error information
                return {
                    "error": ERROR_WRONG_CODE,
                    "invalid_codes": list(code_ets_values),
                    "available_aliases": org_unit_aliases,
                    "org_unit_name": org_unit.name,
                }

    # Validate period against selected period
    if data:  # Only validate if there's data
        # Get all unique period values from the data
        file_periods = {row["period"] for row in data if row["period"]}

        # Get the selected period from the import (format: YYYY-MM)
        selected_period_yyyymm = the_import.month

        # Convert selected period to "Mon-YY" format for comparison

        expected_file_period = convert_yyyymm_to_file_format(selected_period_yyyymm)

        # Check if any period in the file doesn't match the expected period
        for file_period in file_periods:
            file_period_str = str(file_period).strip()

            if file_period_str != expected_file_period:
                print(
                    f"Period validation failed: '{file_period}' in file doesn't match expected period: '{expected_file_period}'"
                )
                # Return detailed error information
                return {
                    "error": ERROR_WRONG_PERIOD,
                    "file_periods": list(file_periods),
                    "selected_period_yyyymm": selected_period_yyyymm,
                    "expected_file_period": expected_file_period,
                    "org_unit_name": org_unit.name,
                }

    # todo this will need to be updated to use batched queries
    for row in data:
        active = not (
            row["art_stoppage"] or row["death"]
        )  # will need to decide what to do with row["served_elsewhere"] row["transfer_out"]
        next_dispensation_date = row["last_dispensation_date"] + timedelta(days=row["days_dispensed"])

        patient, created = Patient.objects.get_or_create(identifier_code=row["identifier_code"])
        new_period = row["period"]
        new_period = convert_file_format_to_yyyymm(new_period)
        record = Record(
            number=row["number"],
            region=row["region"],
            district=row["district"],
            code_ets=row["code_ets"],
            facility_name=row["facility_name"],
            sex=row["sex"],
            age=row["age"],
            weight=row["weight"],
            new_inclusion=row["new_inclusion"],
            transfer_in=row["transfer_in"],
            return_to_care=row["return_to_care"],
            tb_hiv=row["tb_hiv"],
            hiv_type=row["hiv_type"],
            treatment_line=row["treatment_line"],
            last_dispensation_date=row["last_dispensation_date"],
            days_dispensed=row["days_dispensed"],
            next_dispensation_date=next_dispensation_date,
            regimen=row["regimen"],
            stable=row["stable"],
            transfer_out=row["transfer_out"],
            death=row["death"],
            art_stoppage=row["art_stoppage"],
            served_elsewhere=row["served_elsewhere"],
            import_source=the_import,
            org_unit=the_import.org_unit,
            patient=patient,
            period=new_period,
        )
        record.save()

        if created:
            patient.active = active
            patient.last_record = record
            print("Creating patient %s" % patient.identifier_code, "record", record.id)
        else:
            print(new_period, type(new_period))
            if not (patient.last_record) or (new_period > patient.last_record.period[:7]):
                patient.active = active
                print("Updating patient %s" % patient.identifier_code, "record", record.id)
                patient.last_record = record
        patient.save()
        patient.refresh_from_db()
        patient.evaluate_loss(save=True)
