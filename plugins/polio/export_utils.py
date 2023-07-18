import calendar
import datetime as dt
from typing import Any, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

CALENDAR_COLUMNS_CELL_WIDTH = 25.75
CALENDAR_FIRST_COLUMN_CELL_WIDTH = 35.00
CALENDAR_FIRST_COLUMN_CELL_HEIGHT = 40.75
CALENDAR_CELL_WIDTH = 22.00
CALENDAR_CELL_HEIGHT = 40.00
CALENDAR_COLUMN_FONT_SIZE = 12
CALENDAR_CELL_FONT_SIZE = 10


def generate_xlsx_campaigns_calendar(filename: str, datas: Any) -> Workbook:
    """
    Create the XLSX file for the campaigns calendar

        Parameters:
                filename (str): a filename String
                datas (list[dict]): a list of data dictionaries

        returns:
                file (openpyxl.workbook): Saved file openpyxl.workbook object
    """
    file = Workbook()
    sheet = file.active
    sheet.title = filename
    columns = get_columns_names()
    # display columns in the xlsx file
    for column in range(1, len(columns) + 1):
        cell_header = sheet.cell(column=column, row=1, value=columns[column - 1])
        cell_header = font_alignment(cell_header, CALENDAR_COLUMN_FONT_SIZE, "center")
        cell_header = cell_border(cell_header)
        sheet = cell_dimension_pattern_fill(sheet, cell_header, None, CALENDAR_COLUMNS_CELL_WIDTH, True)
    # display calendar data in the xlsx file by looping over each row representing a country campaign rounds
    last_end_row = 0
    for row in range(1, len(datas) + 1):
        # merge cells depending on the max length for rounds it has
        max_rounds_count = get_max_rounds_count(datas[row - 1]["rounds"])
        # specify from which cell to start merging until to which cell
        start_row = row + 1 if row == 1 else last_end_row + 1
        end_row = row + max_rounds_count if row == 1 else last_end_row + max_rounds_count
        # merging cells
        sheet.merge_cells(start_column=1, end_column=1, start_row=start_row, end_row=end_row)
        cell_country = sheet.cell(row=start_row, column=1)
        # assign value to merged cell
        cell_country.value = datas[row - 1]["country_name"]
        cell_country = font_alignment(cell_country, CALENDAR_COLUMN_FONT_SIZE)
        sheet = cell_dimension_pattern_fill(
            sheet, cell_country, CALENDAR_FIRST_COLUMN_CELL_WIDTH, CALENDAR_FIRST_COLUMN_CELL_HEIGHT, True
        )

        last_end_row = end_row
        cell_country_format = sheet.cell(column=1, row=last_end_row)
        cell_border(cell_country_format, False, True)
        # loop over rounds for each month
        for month in range(1, 13):
            if str(month) in datas[row - 1]["rounds"].keys():
                # assign rounds to each cell
                for r in range(0, len(datas[row - 1]["rounds"][str(month)])):
                    cell_val = datas[row - 1]["rounds"][str(month)][r]
                    formatted_cell_val = get_cell_data(cell_val)
                    cell = sheet.cell(column=month + 1, row=r + start_row, value=formatted_cell_val)
                    cell = font_alignment(cell, CALENDAR_CELL_FONT_SIZE)
                    cell = cell_border(cell, True)
                    vaccines_color = None
                    if datas[row - 1]["rounds"][str(month)][r]["vaccines"] != "":
                        vaccines_color = polio_vaccines(datas[row - 1]["rounds"][str(month)][r]["vaccines"])
                    sheet = cell_dimension_pattern_fill(
                        sheet, cell, CALENDAR_CELL_WIDTH, CALENDAR_CELL_HEIGHT, True, vaccines_color
                    )

            # format the last cell in the month column according it has value or not
            cell_format = sheet.cell(column=month + 1, row=last_end_row)
            if cell_format.value is None or cell_format.value == "":
                cell_border(cell_format, False, True)
            else:
                cell_border(cell_format, False, False)

    file.save(filename)

    return file


def get_max_rounds_count(rounds: Any) -> int:
    """
    returns the max length for country rounds

        parameters:
            rounds : all rounds owned by the country
        returns:
            max (int): the max length for all rounds(in each month)
    """
    rounds_list = []
    for key, round in rounds.items():
        rounds_list.append(len(round))
    return max(rounds_list)


def get_cell_data(round: Any) -> str:
    """
    Returns a value(string of one round in a month) to be assigned in an XLSX cell

            parameters:
                round (dict): a round dict

            returns:
                cell_data (string): a cell_data string
    """
    started_at = format_date(round["started_at"])
    ended_at = format_date(round["ended_at"], True)
    obr_name = round["obr_name"] if round["obr_name"] is not None else ""
    round_number = round["round_number"] if round["round_number"] is not None else ""
    cell_data = obr_name + "\n"
    cell_data += "Round " + str(round_number) + "\n"
    cell_data += "Dates: " + started_at + " - " + ended_at + "\n"
    cell_data += round["vaccines"] + "\n" if round["vaccines"] is not None else ""
    cell_data += "Target population: " + str(round["target_population"]) + "\n" if round["target_population"] else ""
    cell_data += (
        "Covered target population: " + str(round["percentage_covered_target_population"]) + "%\n"
        if round["percentage_covered_target_population"]
        else ""
    )
    cell_data += "Geographic scope: " + round["nid_or_snid"] if round["nid_or_snid"] else ""

    return cell_data


def format_date(date: str, with_year: bool = False) -> str:
    """
    Returns a formatted date into "%d %B" or "%d %B %Y" format

            parameters:
                date (str): a date string
                with_year (bool): a with_year boolean
            returns:
                formatted_date (str): a formatted_date string
    """
    date_format = "%d %B"
    if with_year:
        date_format += " %Y"
    # There are some legacy campaigns which have rounds without started_at and ended_at date(None)
    # This check(if date is None) helps to avoid to format a None date
    if date is None:
        formatted_date = ""
    else:
        formatted_date = dt.datetime.strptime(date, "%Y-%m-%d").strftime(date_format)
    return formatted_date


def polio_vaccines(vaccine: str) -> Optional[str]:
    """
    returns a vaccine color

        parameters:
            vaccine (str): a vaccine string
        returns:
            None when the color vaccine doesn't exist or the matching color when the vaccine exists
    """
    vaccine_color = {"nOPV2": "00b0f0", "mOPV2": "66ff66", "bOPV": "ffff00"}
    if vaccine not in vaccine_color.keys():
        return None

    return vaccine_color[vaccine]


def cell_dimension_pattern_fill(
    sheet: Any,
    cell: Any,
    width: Union[None, float],
    height: Union[None, float],
    pattern_fill: Optional[bool] = False,
    color: Optional[str] = "999791",
) -> Any:
    """
    Return the openpyxl.worksheet object after applying fill color on a cell

            parameters:
                sheet (openpyxl.worksheet): a sheet openpyxl.worksheet object
                cell (openpyxl.cell): a cell openpyxl.cell object
                width (float): a width float
                height (float): a height float
                pattern_fill (boolean): a pattern_fill boolean with false as default value

            returns:
                sheet (openpyxl.worksheet): a sheet openpyxl.worksheet object
    """
    if width is not None:
        sheet.column_dimensions[cell.column_letter].width = float(width)
    if height is not None:
        sheet.row_dimensions[cell.row].height = float(height)
    if pattern_fill and color is not None:
        sheet[cell.column_letter + str(cell.row)].fill = PatternFill("solid", start_color=color)

    return sheet


def font_alignment(cell: Any, size: int, horizontal: str = "left") -> Any:
    """
    returns a cell already aligned and formatted with font size

        parameters:
            cell (openpyxl.cell): a cell openpyxl.cell object
            size (str): a size string
        returns:
            cell (openpyxl.cell): a sheet openpyxl.cell object
    """
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)
    cell.font = Font(size=size)
    return cell


def cell_border(cell: Any, all: bool = True, bottom_only: bool = False) -> Any:
    """
    Return the openpyxl.cell object after applying  border

            parameters:
                cell (openpyxl.cell): a cell openpyxl.cell object
                all (boolean): all border are thin
                bottom_only(boolean): Anly bottom border
            returns:
                cell (openpyxl.cell): a sheet openpyxl.cell object
    """
    if all:
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
    elif bottom_only:
        border = Border(
            bottom=Side(style="medium"),
        )
    else:
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="medium"),
        )
    cell.border = border
    return cell


def get_columns_names() -> list:
    """
    Return the list of columns name(COUNTRY and year's months)

            parameters:
                No parameters

            returns:
                columns_names (list): a columns_names list
    """
    columns_names = []
    for month_num in range(1, 13):
        month_name = calendar.month_name[month_num]
        columns_names.append(month_name)
    columns_names.insert(0, "COUNTRY")
    return columns_names


def xlsx_file_name(name: str, params: Any) -> str:
    """
    Return the exported XLSX file name

            parameters:
                name (str): a default name(calendar) string
                params (django.http.request.QueryDict): django query dict params
            returns:
                filename (str): filename string
    """
    current_date = params.get("currentDate")
    campaign_type = params.get("campaignType")
    filename = name
    filename += "_" + current_date if current_date is not None else ""
    filename += "_" + campaign_type if campaign_type is not None else ""
    filename += "_" + "_".join(params.get("countries").split(",")) if params.get("countries") is not None else ""
    filename += (
        "_" + "_".join(params.get("campaignGroups").split(",")) if params.get("campaignGroups") is not None else ""
    )
    return filename
