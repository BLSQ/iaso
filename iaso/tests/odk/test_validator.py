import io
import json

import openpyxl
import pandas as pd

from django.test import SimpleTestCase

from iaso.odk.validator import get_list_name_from_select, parse_sheet, validate_xls_form


class ValidatorTestCase(SimpleTestCase):
    def test_parse_xls_form_valid(self):
        with open("iaso/tests/fixtures/odk_instance_repeat_group_form.xlsx", "rb") as xls_file:
            errors = validate_xls_form(xls_file)
            self.assertEqual(errors, [])

    def test_parse_xls_form_invalid(self):
        with open("iaso/tests/fixtures/odk_invalid_xlsform_expected_errors.json") as expected_errors_file:
            expected_errors = json.loads(expected_errors_file.read())
        with open("iaso/tests/fixtures/odk_invalid_xlsform.xlsx", "rb") as xls_file:
            errors = validate_xls_form(xls_file)
            self.assertEqual(errors, expected_errors)

    def test_parse_sheet_strips_header_trailing_blanks(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "survey"
        ws.append(["type", "name ", "label"])
        ws.append(["text", "q1", "Question 1"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        excel_file = pd.ExcelFile(buf, engine="openpyxl")
        rows = parse_sheet(excel_file, "survey")

        self.assertGreater(len(rows), 0)
        self.assertIn("name", rows[0])
        self.assertNotIn("name ", rows[0])

    def _make_xlsx_buf(self, survey_rows, choices_rows=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "survey"
        for row in survey_rows:
            ws.append(row)
        if choices_rows:
            wc = wb.create_sheet("choices")
            for row in choices_rows:
                wc.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "form.xlsx"
        return buf

    def test_invalid_parameters_column(self):
        buf = self._make_xlsx_buf(
            [
                ["type", "name", "label", "parameters"],
                ["text", "q1", "Question 1", "nd_period"],
            ]
        )
        errors = validate_xls_form(buf)
        self.assertEqual(len(errors), 1)
        self.assertIn("nd_period", errors[0]["message"])
        self.assertEqual(errors[0]["severity"], "error")
        self.assertEqual(errors[0]["sheet"], "survey")

    def test_valid_parameters_column(self):
        buf = self._make_xlsx_buf(
            [
                ["type", "name", "label", "parameters"],
                ["text", "q1", "Question 1", "randomize=true seed=42"],
            ]
        )
        errors = validate_xls_form(buf)
        self.assertEqual(errors, [])

    def test_valid_parameters_column_range(self):
        buf = self._make_xlsx_buf(
            [
                ["type", "name", "label", "parameters"],
                ["range", "amount", "What is the age of the child?", "start=0 end=17 step=1"],
            ]
        )
        errors = validate_xls_form(buf)
        self.assertEqual(errors, [])

    def test_invalid_parameters_column_multiple_equals(self):
        buf = self._make_xlsx_buf(
            [
                ["type", "name", "label", "parameters"],
                ["range", "amount", "Age", "start=0=extra end=17 step=1"],
            ]
        )
        errors = validate_xls_form(buf)
        self.assertEqual(len(errors), 1)
        self.assertIn("start=0=extra", errors[0]["message"])
        self.assertEqual(errors[0]["severity"], "error")

    def test_valid_parameters_column_multiline(self):
        buf = self._make_xlsx_buf(
            [
                ["type", "name", "label", "parameters"],
                ["range", "amount", "What is the age of the child?", "start=0\nend=17\nstep=1"],
            ]
        )
        errors = validate_xls_form(buf)
        self.assertEqual(errors, [])

    def test_get_list_name_from_select_options(self):
        self.assertEqual(get_list_name_from_select({"type": "select_one demo_choices or_other"}), "demo_choices")

        self.assertEqual(get_list_name_from_select({"type": "select_one demo_choices "}), "demo_choices")

        self.assertEqual(get_list_name_from_select({"type": "select_one  demo_choices "}), "demo_choices")

        self.assertEqual(get_list_name_from_select({"type": "select one  demo_choices "}), "demo_choices")

        self.assertEqual(get_list_name_from_select({"type": "select one demo_choices"}), "demo_choices")

        self.assertEqual(get_list_name_from_select({"type": "select one demo_choices or_other "}), "demo_choices")
