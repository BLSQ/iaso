"""Packaged Org_Units.xlsx: a generic list of org units."""

from __future__ import annotations

from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


_PACKAGE_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX_PATH = _PACKAGE_DIR / "Org_Units.xlsx"

SHEET_README = "readme"
SHEET_TYPES = "org_unit_types"

SAMPLE_TYPES: list[dict[str, Any]] = [
    {"name": "Country", "short_name": "COUNTRY", "depth": 1, "parent_type": ""},
    {"name": "Region", "short_name": "REG", "depth": 2, "parent_type": "Country"},
    {"name": "District", "short_name": "DIST", "depth": 3, "parent_type": "Region"},
    {
        "name": "Health Facility",
        "short_name": "HF",
        "depth": 4,
        "parent_type": "District",
    },
    {
        "name": "Warehouse",
        "short_name": "WH",
        "depth": 4,
        "parent_type": "District",
    },
]

# One sheet per leaf type. Parent columns are shared; last column is the leaf.
SAMPLE_SHEETS: dict[str, list[dict[str, str]]] = {
    "Health Facilities": [
        {
            "Country": "Demo Republic",
            "Region": "North",
            "District": "Lake",
            "Health Facility": "North Lake Clinic",
        },
        {
            "Country": "Demo Republic",
            "Region": "North",
            "District": "Lake",
            "Health Facility": "North Lake Hospital",
        },
        {
            "Country": "Demo Republic",
            "Region": "North",
            "District": "Hill",
            "Health Facility": "Hill Health Post",
        },
        {
            "Country": "Demo Republic",
            "Region": "South",
            "District": "Coast",
            "Health Facility": "Coast Clinic",
        },
        {
            "Country": "Demo Republic",
            "Region": "South",
            "District": "Coast",
            "Health Facility": "Port Hospital",
        },
        {
            "Country": "Demo Republic",
            "Region": "South",
            "District": "Valley",
            "Health Facility": "Valley Health Centre",
        },
    ],
    "Warehouses": [
        {
            "Country": "Demo Republic",
            "Region": "North",
            "District": "Lake",
            "Warehouse": "North Depot",
        },
        {
            "Country": "Demo Republic",
            "Region": "South",
            "District": "Coast",
            "Warehouse": "South Depot",
        },
    ],
}

_README = (
    "Create a geographic pyramid from a list of org units.\n"
    "\n"
    "Data sheets: one row per leaf org unit. Columns are the hierarchy from "
    "left to right (parent → child). The last column is the leaf type for "
    "that sheet. Parent columns must be the same on every data sheet.\n"
    "\n"
    "GEO_LEVELS are the shared parent columns. LEAF_TYPES are the last column "
    "of each data sheet. Optional org_unit_types sets short_name / depth / "
    "parent_type. Sheets named readme, org_unit_types, or types are ignored.\n"
    "\n"
    "This sample is fake (Demo Republic). Replace the sheets with any pyramid."
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_KEY_FILL = PatternFill("solid", fgColor="D6E3F0")


def _write_header(sheet: Worksheet, values: list[str]) -> None:
    sheet.append(values)
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 22


def _autosize(sheet: Worksheet, *, min_width: int = 14, max_width: int = 40) -> None:
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        length = min_width
        for cell in column:
            length = max(length, min(max_width, len(str(cell.value or "")) + 2))
        sheet.column_dimensions[letter].width = length


def _regular_sheet(workbook: Workbook, title: str | None = None) -> Worksheet:
    sheet = workbook.active if title is None else workbook.create_sheet(title)
    return cast(Worksheet, sheet)


def write_sample_workbook(path: Path | None = None) -> Path:
    """Write the packaged fake org-unit list (Demo Republic)."""
    target = Path(path) if path is not None else DEFAULT_XLSX_PATH
    workbook = Workbook()

    readme = _regular_sheet(workbook)
    readme.title = SHEET_README
    readme["A1"] = _README
    readme["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    readme.column_dimensions["A"].width = 100
    readme.row_dimensions[1].height = 140
    readme.sheet_view.showGridLines = False

    types_sheet = _regular_sheet(workbook, SHEET_TYPES)
    _write_header(types_sheet, ["name", "short_name", "depth", "parent_type"])
    for row in SAMPLE_TYPES:
        types_sheet.append([row["name"], row["short_name"], row["depth"], row["parent_type"]])
    for row_index in range(2, types_sheet.max_row + 1):
        types_sheet.cell(row_index, 1).fill = _KEY_FILL
    _autosize(types_sheet)

    for sheet_name, rows in SAMPLE_SHEETS.items():
        sheet = _regular_sheet(workbook, sheet_name)
        headers = list(rows[0].keys())
        _write_header(sheet, headers)
        for row in rows:
            sheet.append([row[name] for name in headers])
        _autosize(sheet)

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target
