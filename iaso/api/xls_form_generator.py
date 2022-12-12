import os
import uuid

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend  # type: ignore
from rest_framework.decorators import action
from rest_framework import serializers, filters
from rest_framework.viewsets import ModelViewSet
import string
from datetime import datetime
from tempfile import NamedTemporaryFile
import openpyxl
from iaso.api.common import DeletionFilterBackend, TimestampField, CONTENT_TYPE_XLSX
from iaso.models import Group
from iaso.models.xls_form_template import XlsFormTemplate


polio_plugin = os.environ.get("polio")
if polio_plugin:
    from plugins.polio.models import Campaign


class XlsFormTemplateSerializer(serializers.ModelSerializer):
    class Meta:
        model = XlsFormTemplate
        fields = "__all__"

    created_at = TimestampField(read_only=True)
    updated_at = TimestampField(read_only=True)


def get_data_from_campaigns(campaign_id: uuid, row: any, q_sheet: any, calculation_index: int):
    """This function get the data from a campaign and insert it in the calculation row
    of the xlsform. It's tagged as 'insert_field' in the xlsform.
    """
    authorized_fields = [
        "id",
        "epid",
        "obr_name",
        "gpei_email",
        "description",
        "creation_email_send_at",
        "onset_at",
        "three_level_call_at",
        "cvdpv_notified_at",
        "cvdpv2_notified_at",
        "pv_notified_at",
        "pv2_notified_at",
        "virus",
        "detection_status",
        "detection_responsible",
        "detection_first_draft_submitted_at",
        "detection_rrt_oprtt_approval_at",
        "risk_assessment_status",
        "risk_assessment_responsible",
        "investigation_at",
        "risk_assessment_first_draft_submitted_at",
        "risk_assessment_rrt_oprtt_approval_at",
        "ag_nopv_group_met_at",
        "dg_authorized_at",
        "verification_score",
        "doses_requested",
        "preperadness_spreadsheet_url",
        "preperadness_sync_status",
        "surge_spreadsheet_url",
        "country_name_in_surge_spreadsheet",
        "budget_status",
        "budget_responsible",
        "created_at",
        "updated_at",
        "district_count",
        "round_one",
        "round_two",
        "vacine",
        "obr_name",
    ]
    campaign = get_object_or_404(Campaign, id=campaign_id)
    for i in range(2, row + 1):
        cell_obj = q_sheet.cell(row=i, column=2)
        cell_value_start = cell_obj.value[:7] if cell_obj.value is not None else ""
        if cell_value_start == "insert_":
            str_request = cell_obj.value[7:]
            if str_request in authorized_fields:
                cell_obj = q_sheet.cell(row=i, column=calculation_index)
                cell_obj.value = str(getattr(campaign, str_request))


def get_column_position(column_name, sheet):
    for row_calc in sheet.rows:
        column_position = 0
        for s_cell in row_calc:
            column_position += 1
            if str(s_cell.value).lower() == column_name:
                return column_position


def create_ou_tree_list(group_ou):
    """
    This function create a list of dictionaries that contains orgunits from top of hierarchy to bottom.
    """
    # create list of dictionary with OU tree
    ou_tree_list = []
    for ou in group_ou:
        ou_tree_dict = {ou.org_unit_type.name: ou for ou in ou.create_ou_tree()}
        ou_tree_list.append(ou_tree_dict)
    return ou_tree_list


class XlsFormGeneratorViewSet(ModelViewSet):
    results_key = "xlsformgenerator"
    remove_results_key_if_paginated = True
    filter_backends = [filters.OrderingFilter, DjangoFilterBackend, DeletionFilterBackend]

    def get_serializer_class(self):
        return XlsFormTemplateSerializer

    def get_queryset(self):
        queryset = XlsFormTemplate.objects.filter(account=self.request.user.iaso_profile.account)
        return queryset

    @action(methods=["GET"], detail=False)
    def xlsform_generator(self, request) -> HttpResponse:
        """
        Export a xlsform by using a group as OU source. A template is required. Specific data can be extracted from campaigns by
        starting the name variable by 'insert_' in the xls file in "calculate" row. The data will be saved in the
        calculation column. By adding campaignid to the params you can extract data from a campaign.
        """
        group_id = request.query_params.get("groupid", None)
        form_name = request.query_params.get("form_name", None)
        campaign_id = request.query_params.get("campaignid", None)
        org_unit_type_list = request.GET.get("outypelist", None).split(",")
        ou_hierarchy_list = request.GET.get("ouhierarchy", None).split(",")

        if not form_name:
            raise serializers.ValidationError({"error": "No form provided."})

        group = get_object_or_404(Group, id=group_id)

        # Check if User has access to Group
        if not group.user_has_access_to(request.user):
            raise serializers.ValidationError({"error": "You don't have the access to this Org Unit group."})

        if not ou_hierarchy_list:
            raise serializers.ValidationError({"error": "You must provied an Org unit hierarchy."})

        try:
            path = XlsFormTemplate.objects.get(name=form_name).form_template.path
        except ValueError:
            raise serializers.ValidationError({"error": f"Bad Template Name."})

        group_ou = group.org_units.all()

        ou_tree_list = create_ou_tree_list(group_ou)

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        sheet = wb.get_sheet_by_name("choices")  # type: ignore
        choices_row = 2
        choices_column = 1
        cell = list(string.ascii_uppercase)
        q_sheet = wb.get_sheet_by_name("survey")  # type: ignore
        survey_columns = []
        survey_last_empty_row = len(list(q_sheet.rows))
        for l in cell:
            survey_columns.append(q_sheet[f"{l}1"].value)

        # create xls columns
        sheet[cell[choices_column - 1] + "1"] = "list name"
        sheet[cell[choices_column] + "1"] = "name"
        sheet[cell[choices_column + 1] + "1"] = "label"
        sheet[cell[choices_column + 2] + "1"] = "id"
        i = 2
        for ou_type in org_unit_type_list:
            sheet[cell[choices_column + i] + "1"] = ou_type
            i += 1

        # insert rows to add the org units fields at the top of the file
        ws.insert_rows(3, 5)

        added_ou_list = []
        key_added_to_choice_list = []
        starting_row = 3
        ou_hierarchy_list = [d.lower() for d in ou_hierarchy_list]

        # populate xls with OU
        for ou_dic in ou_tree_list:
            for k, v in ou_dic.items():
                if v.id not in added_ou_list:
                    added_ou_list.append(v.id)
                    if k not in key_added_to_choice_list:
                        # Add the ou type to the select choices of the xls form
                        key_added_to_choice_list.append(k)
                        q_sheet[
                            cell[get_column_position("type", q_sheet) - 1] + str(starting_row)
                        ] = f"select_one ou_{str(k).lower()}"
                        q_sheet[
                            cell[get_column_position("name", q_sheet) - 1] + str(starting_row)
                        ] = f"ou_{str(k).lower()}"
                        q_sheet[cell[get_column_position("label", q_sheet) - 1] + str(starting_row)] = f"Select {k}"
                        q_sheet[cell[get_column_position("required", q_sheet) - 1] + str(starting_row)] = "yes"
                        if ou_hierarchy_list.index(k.lower()) != 0:
                            parent_type = ou_hierarchy_list[ou_hierarchy_list.index(str(k.lower())) - 1]
                            q_sheet[cell[get_column_position("choice_filter", q_sheet) - 1] + str(starting_row)] = (
                                f"{parent_type}=$" "{" f"ou_{parent_type}" "} "
                            )
                    # Add org units as hierarchy into the choices sheet.
                    sheet[cell[choices_column - 1] + str(choices_row)] = f"ou_{str(k).lower()}"
                    sheet[cell[choices_column] + str(choices_row)] = v.id
                    sheet[cell[choices_column + 1] + str(choices_row)] = str(v.name)

                    index_hierarchy = ou_hierarchy_list.index(str(k.lower()))
                    ou_dic = {k.lower(): v for k, v in ou_dic.items()}

                    calculation_index = 0
                    if index_hierarchy > 0:
                        index_hierarchy -= 1
                        for row_calc in sheet.rows:
                            iterator = 0
                            for s_cell in row_calc:
                                iterator += 1
                                if str(s_cell.value).lower() == ou_hierarchy_list[index_hierarchy].lower():
                                    calculation_index += iterator
                                    break
                    parent_ou_id = ou_dic.get(ou_hierarchy_list[index_hierarchy].lower()).id
                    sheet[cell[choices_column + calculation_index - 2] + str(choices_row)] = str(parent_ou_id)

                    choices_row += 1
                    starting_row += 1
                    survey_last_empty_row += 1

        row = q_sheet.max_row
        calculation_index = get_column_position("calculation", q_sheet)

        # Insert data as calculation from campaigns
        if campaign_id and request.user.has_perm("iaso_polio") and polio_plugin:
            get_data_from_campaigns(campaign_id, row, q_sheet, calculation_index)

        filename = f"FORM_{form_name}_{datetime.now().date()}.xlsx"

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

            response = HttpResponse(stream, content_type=CONTENT_TYPE_XLSX)
            response["Content-Disposition"] = "attachment; filename=%s" % filename
            return response
