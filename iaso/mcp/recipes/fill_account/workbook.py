from __future__ import annotations

import json

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet


_PACKAGE_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX_PATH = _PACKAGE_DIR / "account_setup_template.xlsx"

SHEET_README = "readme"
SHEET_SETTINGS = "settings"
SHEET_TYPES = "org_unit_types"
SHEET_ORG_UNITS = "org_units"
SHEET_GEO = "org_unit_geo"
SHEET_FORMS = "forms"
SHEET_ENTITY_TYPES = "entity_types"
SHEET_WORKFLOWS = "workflows"
SHEET_FOLLOWUPS = "workflow_followups"
SHEET_CHANGES = "workflow_changes"
SHEET_ENTITIES = "entities"
SHEET_INSTANCES = "instances"

SETTING_KEYS = (
    "project_name",
    "project_app_id",
    "project_description",
    "project_color",
    "data_source_name",
    "source_version_description",
    "set_as_default",
)

SAMPLE_SETTINGS: dict[str, str] = {
    "project_name": "Demo",
    "project_app_id": "org.bluesquare.iaso.demo",
    "project_description": "Prospect demo project",
    "project_color": "#1976D2",
    "data_source_name": "Demo pyramid",
    "source_version_description": "Sample health pyramid for prospect demos",
    "set_as_default": "true",
}

SAMPLE_TYPES: list[dict[str, Any]] = [
    {
        "name": "Country",
        "short_name": "COUNTRY",
        "depth": 1,
        "parent_type": "",
    },
    {
        "name": "Region",
        "short_name": "REG",
        "depth": 2,
        "parent_type": "Country",
    },
    {
        "name": "District",
        "short_name": "DIST",
        "depth": 3,
        "parent_type": "Region",
    },
    {
        "name": "Health Facility",
        "short_name": "HF",
        "depth": 4,
        "parent_type": "District",
    },
]

SAMPLE_ORG_UNITS: list[dict[str, str]] = [
    {
        "Country": "Demo Republic",
        "Region": "North",
        "District": "Lake",
        "Health Facility": "North Lake Clinic",
    },
    {
        "Country": "Demo Republic",
        "Region": "North",
        "District": "Lake",
        "Health Facility": "North Lake Hospital",
    },
    {
        "Country": "Demo Republic",
        "Region": "North",
        "District": "Hill",
        "Health Facility": "Hill Health Post",
    },
    {
        "Country": "Demo Republic",
        "Region": "South",
        "District": "Coast",
        "Health Facility": "Coast Clinic",
    },
    {
        "Country": "Demo Republic",
        "Region": "South",
        "District": "Coast",
        "Health Facility": "Port Hospital",
    },
    {
        "Country": "Demo Republic",
        "Region": "South",
        "District": "Valley",
        "Health Facility": "Valley Health Centre",
    },
]

# Nested fake geography (lon/lat). Admin units are polygons; facilities are GPS points.
# bbox is west,south,east,north so Max/Laure can edit boxes without GeoJSON.
SAMPLE_ORG_UNIT_GEO: list[dict[str, Any]] = [
    {"name": "Demo Republic", "bbox": "14,11,16,13"},
    {"name": "North", "bbox": "14,12,16,13"},
    {"name": "South", "bbox": "14,11,16,12"},
    {"name": "Lake", "bbox": "14,12,15,13"},
    {"name": "Hill", "bbox": "15,12,16,13"},
    {"name": "Coast", "bbox": "14,11,15,12"},
    {"name": "Valley", "bbox": "15,11,16,12"},
    {"name": "North Lake Clinic", "latitude": 12.42, "longitude": 14.32},
    {"name": "North Lake Hospital", "latitude": 12.71, "longitude": 14.68},
    {"name": "Hill Health Post", "latitude": 12.38, "longitude": 15.45},
    {"name": "Coast Clinic", "latitude": 11.35, "longitude": 14.28},
    {"name": "Port Hospital", "latitude": 11.62, "longitude": 14.72},
    {"name": "Valley Health Centre", "latitude": 11.44, "longitude": 15.51},
]

SAMPLE_FORMS: list[dict[str, str]] = [
    {
        "name": "Beneficiary registration",
        "form_id": "demo_beneficiary_registration",
        "org_unit_type": "Health Facility",
        "kind": "profile",
    },
    {
        "name": "Beneficiary visit",
        "form_id": "demo_beneficiary_visit",
        "org_unit_type": "Health Facility",
        "kind": "visit",
    },
    {
        "name": "Beneficiary update",
        "form_id": "demo_beneficiary_update",
        "org_unit_type": "Health Facility",
        "kind": "update",
    },
]

SAMPLE_ENTITY_TYPES: list[dict[str, str]] = [
    {
        "name": "Beneficiary",
        "reference_form": "Beneficiary registration",
        "list_fields": "first_name,last_name,sex,age",
        "detail_fields": "first_name,last_name,sex,age",
        "duplicate_fields": "first_name,last_name",
    },
]

SAMPLE_ENTITIES: list[dict[str, str]] = [
    {
        "name": "Ada Ndiaye",
        "entity_type": "Beneficiary",
        "org_unit": "North Lake Clinic",
        "first_name": "Ada",
        "last_name": "Ndiaye",
        "sex": "F",
        "age": "28",
    },
    {
        "name": "Omar Diallo",
        "entity_type": "Beneficiary",
        "org_unit": "Coast Clinic",
        "first_name": "Omar",
        "last_name": "Diallo",
        "sex": "M",
        "age": "34",
    },
    {
        "name": "Fatou Sow",
        "entity_type": "Beneficiary",
        "org_unit": "Hill Health Post",
        "first_name": "Fatou",
        "last_name": "Sow",
        "sex": "F",
        "age": "22",
    },
    {
        "name": "Jean Pierre",
        "entity_type": "Beneficiary",
        "org_unit": "Valley Health Centre",
        "first_name": "Jean",
        "last_name": "Pierre",
        "sex": "M",
        "age": "41",
    },
]

SAMPLE_INSTANCES: list[dict[str, str]] = [
    {
        "entity": "Ada Ndiaye",
        "form": "Beneficiary visit",
        "org_unit": "North Lake Clinic",
        "visit_date": "2026-03-12",
        "weight_kg": "54.2",
        "muac_mm": "135",
    },
    {
        "entity": "Omar Diallo",
        "form": "Beneficiary visit",
        "org_unit": "Coast Clinic",
        "visit_date": "2026-03-18",
        "weight_kg": "68.0",
        "muac_mm": "142",
    },
    {
        "entity": "Fatou Sow",
        "form": "Beneficiary visit",
        "org_unit": "Hill Health Post",
        "visit_date": "2026-04-02",
        "weight_kg": "49.5",
        "muac_mm": "128",
    },
    {
        "entity": "Jean Pierre",
        "form": "Beneficiary visit",
        "org_unit": "Valley Health Centre",
        "visit_date": "2026-04-09",
        "weight_kg": "72.1",
        "muac_mm": "148",
    },
]

SAMPLE_WORKFLOWS: list[dict[str, str]] = [
    {
        "name": "Beneficiary care",
        "entity_type": "Beneficiary",
        "auto_first_step": "true",
        "publish": "true",
    },
]

SAMPLE_WORKFLOW_FOLLOWUPS: list[dict[str, str]] = [
    {
        "workflow": "Beneficiary care",
        "form": "Beneficiary visit",
        "order": "0",
        "condition": "true",
    },
    {
        "workflow": "Beneficiary care",
        "form": "Beneficiary update",
        "order": "1",
        "condition": "true",
    },
]

SAMPLE_WORKFLOW_CHANGES: list[dict[str, str]] = [
    {
        "workflow": "Beneficiary care",
        "form": "Beneficiary update",
        "mapping": "first_name,last_name,sex,age",
    },
]

_README = """IASO account setup workbook

Use this file to fill an empty (or almost empty) IASO account with demo data
for prospect showcases. Duplicate the template, replace the sample Demo
Republic with country-shaped fake data, then preview and apply.

How to run
1. Duplicate this file (or ask the IASO MCP to write a new one).
2. Edit the yellow cells. Grey keys / notes are guidance — leave the key names as they are.
3. Preview, then apply:
     uv run python -m cases.fill_account_from_xls
   or MCP: preview_fill_account / apply_fill_account (confirm=true).

Sheets
settings         Project + data source. Matched by name on re-run.
org_unit_types   Hierarchy. depth 1 is the root. Column names on org_units must match name.
org_units        One row per facility. Left = parent, right = leaf. Parents are created from unique paths.
org_unit_geo     GPS points (latitude/longitude) and shapes. shape: GeoJSON or bbox west,south,east,north.
forms            XLSForms. kind: profile (entity attributes), visit, or update. form_id is the ODK id.
entity_types     Linked to a profile form. list_fields are shown in the entity list.
workflows        One entity workflow per entity type. publish=true after follow-ups and changes.
workflow_followups  Mobile follow-up forms (condition true = always). workflow matches workflows.name.
workflow_changes Mapping from an update/visit form onto the profile (source=target or identity names).
entities         People: name, entity_type, org_unit, plus extra columns = form answers.
instances        Visits: entity, form, org_unit, plus extra columns = form answers.

Rules
- Fills the IASO account of the current FastMCP session (does not create a new Account tenant).
- Re-runs reuse project / source / types / org units / forms / entity types / published workflows with the same name.
- Entities and submissions use a stable uuid from the name (get-or-create).
- project_app_id must be unique on the IASO instance if the project is new.
- set_as_default=true sets this source as the account default version (needed for the IASO UI).
- Dropdowns on kind / type / org unit help keep names consistent.
- Do not apply until a preview looks right.
"""

_BLUE = "1976D2"
_BLUE_DARK = "0D47A1"
_YELLOW = "FFF8E1"
_YELLOW_ALT = "FFECB3"
_GREY = "F5F5F5"
_GREY_TEXT = "616161"
_WHITE = "FFFFFF"
_ROW_ALT = "F7FBFF"
_LEVEL_HEADER = ("1565C0", "2E7D32", "EF6C00", "6A1B9A", "00838F", "C62828")
_LEVEL_CELL = ("E3F2FD", "E8F5E9", "FFF3E0", "F3E5F5", "E0F7FA", "FFEBEE")
_TAB: dict[str, str] = {
    SHEET_README: "90A4AE",
    SHEET_SETTINGS: "1976D2",
    SHEET_TYPES: "00897B",
    SHEET_ORG_UNITS: "43A047",
    SHEET_GEO: "2E7D32",
    SHEET_FORMS: "7B1FA2",
    SHEET_ENTITY_TYPES: "EF6C00",
    SHEET_WORKFLOWS: "5E35B1",
    SHEET_FOLLOWUPS: "00838F",
    SHEET_CHANGES: "6D4C41",
    SHEET_ENTITIES: "F9A825",
    SHEET_INSTANCES: "EC407A",
}

_TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=_WHITE)
_SECTION_FONT = Font(name="Calibri", size=13, bold=True, color=_BLUE_DARK)
_README_FONT = Font(name="Calibri", size=11, color="212121")
_README_MUTED = Font(name="Calibri", size=11, italic=True, color=_GREY_TEXT)
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=_WHITE)
_KEY_FONT = Font(name="Calibri", size=11, bold=True, color="37474F")
_NOTE_FONT = Font(name="Calibri", size=10, italic=True, color=_GREY_TEXT)
_BODY_FONT = Font(name="Calibri", size=11, color="212121")
_TITLE_FILL = PatternFill("solid", fgColor=_BLUE)
_SECTION_FILL = PatternFill("solid", fgColor="E3F2FD")
_HEADER_FILL = PatternFill("solid", fgColor=_BLUE)
_KEY_FILL = PatternFill("solid", fgColor=_GREY)
_EDIT_FILL = PatternFill("solid", fgColor=_YELLOW)
_EDIT_ALT_FILL = PatternFill("solid", fgColor=_YELLOW_ALT)
_ALT_FILL = PatternFill("solid", fgColor=_ROW_ALT)
_WRAP = Alignment(wrap_text=True, vertical="center")
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
_THIN = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0"),
)
_HEADER_BORDER = Border(
    left=Side(style="thin", color=_BLUE_DARK),
    right=Side(style="thin", color=_BLUE_DARK),
    top=Side(style="thin", color=_BLUE_DARK),
    bottom=Side(style="medium", color=_BLUE_DARK),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _set_text(cell, value: object) -> None:
    """Store as a string so Excel does not treat leading '=' as a formula."""
    text = "" if value is None else str(value)
    cell.value = text
    cell.data_type = "s"


def _append_row(sheet: Worksheet, values: list[object]) -> None:
    cleaned = [None if value == "" else value for value in values]
    sheet.append(cleaned)
    row = sheet.max_row
    for index, value in enumerate(cleaned, start=1):
        if isinstance(value, str):
            _set_text(sheet.cell(row, index), value)


def _header_row(
    sheet: Worksheet,
    values: list[str],
    *,
    comments: dict[str, str] | None = None,
    fills: list[str] | None = None,
) -> None:
    _append_row(sheet, values)
    sheet.row_dimensions[1].height = 28
    for index, name in enumerate(values, start=1):
        cell = sheet.cell(1, index)
        cell.font = _HEADER_FONT
        cell.fill = _fill(fills[index - 1]) if fills and index <= len(fills) else _HEADER_FILL
        cell.alignment = _WRAP
        cell.border = _HEADER_BORDER
        hint = (comments or {}).get(name)
        if hint:
            cell.comment = Comment(hint, "IASO")
            cell.comment.width = 280
            cell.comment.height = 80


def _style_data_rows(
    sheet: Worksheet,
    *,
    editable: bool = True,
    skip_cols: set[int] | None = None,
    level_fills: list[str] | None = None,
) -> None:
    skip_cols = skip_cols or set()
    max_col = sheet.max_column
    for row in range(2, sheet.max_row + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row, col)
            cell.font = _NOTE_FONT if col in skip_cols else _BODY_FONT
            cell.alignment = _WRAP
            cell.border = _THIN
            if col in skip_cols:
                cell.fill = _KEY_FILL
            elif level_fills and col <= len(level_fills):
                cell.fill = _fill(level_fills[col - 1])
            elif editable:
                cell.fill = _EDIT_ALT_FILL if row % 2 == 0 else _EDIT_FILL
            elif row % 2 == 0:
                cell.fill = _ALT_FILL


def _autosize(sheet: Worksheet, *, min_width: int = 14, max_width: int = 48) -> None:
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        length = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, min(max_width, len(value) + 2))
        sheet.column_dimensions[letter].width = max(min_width, length)
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    title = sheet.title or ""
    sheet.sheet_properties.tabColor = _TAB[title] if title in _TAB else _BLUE  # noqa: SIM401
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_view.showGridLines = False


def _dropdown(sheet: Worksheet, formula: str, cell_range: str, prompt: str) -> None:
    validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error="Pick a value from the list (or leave blank).",
        promptTitle="Hint",
        prompt=prompt,
        showInputMessage=True,
    )
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _truthy(value: str) -> bool:
    return value.strip().lower() in {"1", "true", "yes", "y"}


def _split_fields(value: object) -> list[str]:
    text = _cell_str(value)
    if not text:
        return []
    return [part.strip() for part in text.replace(";", ",").split(",") if part.strip()]


def _join_fields(values: list[str] | None) -> str:
    return ",".join(values or [])


def parse_condition(value: object) -> Any:
    text = _cell_str(value)
    if not text or text.lower() in {"true", "always", "yes"}:
        return "true"
    if text.lower() in {"false", "never", "no"}:
        return "false"
    if text[:1] in {"{", "["}:
        return json.loads(text)
    return text


def parse_mapping(value: object) -> dict[str, str]:
    text = _cell_str(value)
    if not text:
        return {}
    if text.startswith("{"):
        parsed = json.loads(text)
        if not isinstance(parsed, dict):
            msg = "mapping JSON must be an object of source → target question names."
            raise ValueError(msg)
        return {str(key): str(val) for key, val in parsed.items()}
    mapping: dict[str, str] = {}
    for part in text.replace(";", ",").split(","):
        piece = part.strip()
        if not piece:
            continue
        if "=" in piece:
            source, target = piece.split("=", 1)
            mapping[source.strip()] = target.strip()
        else:
            mapping[piece] = piece
    return mapping


def _join_mapping(mapping: dict[str, str] | None) -> str:
    if not mapping:
        return ""
    if all(source == target for source, target in mapping.items()):
        return ",".join(mapping)
    return ",".join(f"{source}={target}" for source, target in mapping.items())


def box_shape(west: float, south: float, east: float, north: float) -> dict[str, Any]:
    ring = [
        [west, south],
        [east, south],
        [east, north],
        [west, north],
        [west, south],
    ]
    return {"type": "MultiPolygon", "coordinates": [[ring]]}


def _fmt_coord(value: float) -> str:
    text = f"{value:.6f}".rstrip("0").rstrip(".")
    return text or "0"


def _parse_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _cell_str(value)
    if not text:
        return None
    return float(text)


def _as_multipolygon(geom: dict[str, Any]) -> dict[str, Any]:
    kind = geom.get("type")
    if kind == "FeatureCollection":
        features = geom.get("features") or []
        if not features or not isinstance(features[0], dict):
            msg = "FeatureCollection has no features."
            raise ValueError(msg)
        return _as_multipolygon(features[0].get("geometry") or features[0])
    if kind == "Feature":
        return _as_multipolygon(geom.get("geometry") or {})
    if kind == "Polygon":
        return {"type": "MultiPolygon", "coordinates": [geom["coordinates"]]}
    if kind == "MultiPolygon":
        return {"type": "MultiPolygon", "coordinates": geom["coordinates"]}
    msg = f"shape must be a Polygon or MultiPolygon (got {kind!r})."
    raise ValueError(msg)


def parse_shape(value: object) -> dict[str, Any] | None:
    if value is None or value == "":
        return None
    if isinstance(value, dict):
        return _as_multipolygon(value)
    text = _cell_str(value)
    if not text:
        return None
    if text.startswith(("{", "[")):
        return _as_multipolygon(json.loads(text))
    parts = [part.strip() for part in text.replace(";", ",").split(",") if part.strip()]
    if len(parts) != 4:
        msg = (
            "shape must be GeoJSON or a bbox west,south,east,north "
            "(four numbers: min longitude, min latitude, max longitude, max latitude)."
        )
        raise ValueError(msg)
    west, south, east, north = (float(part) for part in parts)
    if west >= east or south >= north:
        msg = "bbox must have west < east and south < north."
        raise ValueError(msg)
    return box_shape(west, south, east, north)


def _axis_aligned_bbox(shape: dict[str, Any]) -> tuple[float, float, float, float] | None:
    try:
        ring = shape["coordinates"][0][0]
    except (KeyError, IndexError, TypeError):
        return None
    if not ring or ring[0] != ring[-1] or len(ring) != 5:
        return None
    xs = {point[0] for point in ring}
    ys = {point[1] for point in ring}
    if len(xs) != 2 or len(ys) != 2:
        return None
    return (min(xs), min(ys), max(xs), max(ys))


def shape_to_cell(shape: dict[str, Any] | None) -> str:
    if not shape:
        return ""
    bbox = _axis_aligned_bbox(shape)
    if bbox:
        return ",".join(_fmt_coord(value) for value in bbox)
    return json.dumps(shape, separators=(",", ":"))


def sample_workbook() -> dict[str, Any]:
    return {
        "settings": dict(SAMPLE_SETTINGS),
        "org_unit_types": [dict(item) for item in SAMPLE_TYPES],
        "org_units": [dict(row) for row in SAMPLE_ORG_UNITS],
        "org_unit_geo": [dict(row) for row in SAMPLE_ORG_UNIT_GEO],
        "forms": [dict(row) for row in SAMPLE_FORMS],
        "entity_types": [dict(row) for row in SAMPLE_ENTITY_TYPES],
        "entities": [dict(row) for row in SAMPLE_ENTITIES],
        "instances": [dict(row) for row in SAMPLE_INSTANCES],
        "workflows": [dict(row) for row in SAMPLE_WORKFLOWS],
        "workflow_followups": [dict(row) for row in SAMPLE_WORKFLOW_FOLLOWUPS],
        "workflow_changes": [dict(row) for row in SAMPLE_WORKFLOW_CHANGES],
    }


def validate_workbook(data: dict[str, Any]) -> dict[str, Any]:
    settings_raw = data.get("settings") or {}
    if not isinstance(settings_raw, dict):
        msg = "settings must be an object of key/value strings."
        raise ValueError(msg)
    settings = {key: _cell_str(settings_raw.get(key, SAMPLE_SETTINGS.get(key, ""))) for key in SETTING_KEYS}
    missing_settings = [key for key in ("project_name", "project_app_id", "data_source_name") if not settings[key]]
    if missing_settings:
        msg = f"settings missing: {', '.join(missing_settings)}."
        raise ValueError(msg)

    types_raw = data.get("org_unit_types") or []
    if not types_raw:
        msg = "org_unit_types must have at least one row."
        raise ValueError(msg)
    types: list[dict[str, Any]] = []
    by_name: dict[str, dict[str, Any]] = {}
    for index, item in enumerate(types_raw, start=1):
        name = _cell_str(item.get("name"))
        short_name = _cell_str(item.get("short_name")) or name[:12].upper()
        depth_raw = item.get("depth")
        try:
            depth = int(depth_raw)
        except (TypeError, ValueError) as exc:
            msg = f"org_unit_types row {index}: depth must be an integer."
            raise ValueError(msg) from exc
        parent_type = _cell_str(item.get("parent_type"))
        if not name:
            msg = f"org_unit_types row {index}: name is required."
            raise ValueError(msg)
        if name in by_name:
            msg = f"Duplicate org unit type '{name}'."
            raise ValueError(msg)
        row = {
            "name": name,
            "short_name": short_name,
            "depth": depth,
            "parent_type": parent_type,
        }
        types.append(row)
        by_name[name] = row
    types.sort(key=lambda item: item["depth"])
    depths = [item["depth"] for item in types]
    if depths != list(range(1, len(types) + 1)):
        msg = f"org_unit_types depths must be 1..n with no gaps (got {depths})."
        raise ValueError(msg)
    for item in types:
        parent_type = item["parent_type"]
        if item["depth"] == 1:
            if parent_type:
                msg = f"Root type '{item['name']}' must have an empty parent_type."
                raise ValueError(msg)
            continue
        expected = types[item["depth"] - 2]["name"]
        if parent_type and parent_type != expected:
            msg = f"Type '{item['name']}' parent_type '{parent_type}' should be '{expected}'."
            raise ValueError(msg)
        item["parent_type"] = expected

    type_names = [item["name"] for item in types]
    rows_raw = data.get("org_units") or []
    if not rows_raw:
        msg = "org_units must have at least one row."
        raise ValueError(msg)
    rows: list[dict[str, str]] = []
    for index, item in enumerate(rows_raw, start=1):
        row = {name: _cell_str(item.get(name)) for name in type_names}
        if not any(row.values()):
            continue
        empty = [name for name in type_names if not row[name]]
        if empty:
            msg = f"org_units row {index} is missing {empty}. Every row is a full path from country to facility."
            raise ValueError(msg)
        rows.append(row)
    if not rows:
        msg = "org_units has no data rows."
        raise ValueError(msg)

    ou_names = {row[name] for row in rows for name in type_names}

    org_unit_geo: list[dict[str, Any]] = []
    geo_names: set[str] = set()
    for index, item in enumerate(data.get("org_unit_geo") or [], start=1):
        name = _cell_str(item.get("name"))
        if not name:
            continue
        if name in geo_names:
            msg = f"Duplicate org_unit_geo '{name}'."
            raise ValueError(msg)
        if name not in ou_names:
            msg = f"org_unit_geo row {index}: unknown org unit '{name}'."
            raise ValueError(msg)
        try:
            latitude = _parse_float(item.get("latitude"))
            longitude = _parse_float(item.get("longitude"))
            altitude = _parse_float(item.get("altitude"))
        except ValueError as exc:
            msg = f"org_unit_geo row {index}: {exc}"
            raise ValueError(msg) from exc
        if (latitude is None) != (longitude is None):
            msg = f"org_unit_geo row {index}: latitude and longitude must both be set."
            raise ValueError(msg)
        if latitude is not None and not -90 <= latitude <= 90:
            msg = f"org_unit_geo row {index}: latitude must be between -90 and 90."
            raise ValueError(msg)
        if longitude is not None and not -180 <= longitude <= 180:
            msg = f"org_unit_geo row {index}: longitude must be between -180 and 180."
            raise ValueError(msg)
        raw_shape = item.get("shape")
        if raw_shape in (None, "") and item.get("bbox") not in (None, ""):
            raw_shape = item.get("bbox")
        try:
            shape = parse_shape(raw_shape)
        except (json.JSONDecodeError, ValueError, KeyError, TypeError) as exc:
            msg = f"org_unit_geo row {index}: invalid shape ({exc})."
            raise ValueError(msg) from exc
        if latitude is None and shape is None:
            msg = (
                f"org_unit_geo row {index}: set latitude/longitude and/or a shape "
                "(GeoJSON or bbox west,south,east,north)."
            )
            raise ValueError(msg)
        geo_names.add(name)
        org_unit_geo.append(
            {
                "name": name,
                "latitude": latitude,
                "longitude": longitude,
                "altitude": altitude,
                "shape": shape,
            }
        )

    forms: list[dict[str, str]] = []
    form_names: set[str] = set()
    for index, item in enumerate(data.get("forms") or [], start=1):
        name = _cell_str(item.get("name"))
        form_id = _cell_str(item.get("form_id")) or name.lower().replace(" ", "_")
        org_unit_type = _cell_str(item.get("org_unit_type"))
        kind = _cell_str(item.get("kind")) or "form"
        if not name:
            continue
        if name in form_names:
            msg = f"Duplicate form '{name}'."
            raise ValueError(msg)
        if org_unit_type and org_unit_type not in by_name:
            msg = f"forms row {index}: unknown org_unit_type '{org_unit_type}'."
            raise ValueError(msg)
        if kind not in {"profile", "visit", "update", "form"}:
            msg = f"forms row {index}: kind must be profile, visit, update, or form."
            raise ValueError(msg)
        form_names.add(name)
        forms.append(
            {
                "name": name,
                "form_id": form_id,
                "org_unit_type": org_unit_type,
                "kind": kind,
            }
        )

    entity_types: list[dict[str, Any]] = []
    entity_type_names: set[str] = set()
    for index, item in enumerate(data.get("entity_types") or [], start=1):
        name = _cell_str(item.get("name"))
        reference_form = _cell_str(item.get("reference_form"))
        if not name:
            continue
        if name in entity_type_names:
            msg = f"Duplicate entity type '{name}'."
            raise ValueError(msg)
        if reference_form and reference_form not in form_names:
            msg = f"entity_types row {index}: reference_form '{reference_form}' is not in the forms sheet."
            raise ValueError(msg)
        entity_type_names.add(name)
        entity_types.append(
            {
                "name": name,
                "reference_form": reference_form,
                "list_fields": _split_fields(item.get("list_fields")),
                "detail_fields": _split_fields(item.get("detail_fields")),
                "duplicate_fields": _split_fields(item.get("duplicate_fields")),
            }
        )

    entities: list[dict[str, str]] = []
    entity_names: set[str] = set()
    for index, item in enumerate(data.get("entities") or [], start=1):
        record = {key: _cell_str(value) for key, value in item.items()}
        name = record.get("name") or ""
        if not name:
            continue
        if name in entity_names:
            msg = f"Duplicate entity '{name}'."
            raise ValueError(msg)
        entity_type = record.get("entity_type") or ""
        org_unit = record.get("org_unit") or ""
        if entity_type and entity_type not in entity_type_names:
            msg = f"entities row {index}: unknown entity_type '{entity_type}'."
            raise ValueError(msg)
        if org_unit and org_unit not in ou_names:
            msg = f"entities row {index}: unknown org_unit '{org_unit}'."
            raise ValueError(msg)
        entity_names.add(name)
        entities.append(record)

    instances: list[dict[str, str]] = []
    for index, item in enumerate(data.get("instances") or [], start=1):
        record = {key: _cell_str(value) for key, value in item.items()}
        if not any(record.values()):
            continue
        form_name = record.get("form") or ""
        entity_name = record.get("entity") or ""
        org_unit = record.get("org_unit") or ""
        if form_name and form_name not in form_names:
            msg = f"instances row {index}: unknown form '{form_name}'."
            raise ValueError(msg)
        if entity_name and entity_name not in entity_names:
            msg = f"instances row {index}: unknown entity '{entity_name}'."
            raise ValueError(msg)
        if org_unit and org_unit not in ou_names:
            msg = f"instances row {index}: unknown org_unit '{org_unit}'."
            raise ValueError(msg)
        instances.append(record)

    workflows: list[dict[str, Any]] = []
    workflow_names: set[str] = set()
    for index, item in enumerate(data.get("workflows") or [], start=1):
        name = _cell_str(item.get("name"))
        entity_type = _cell_str(item.get("entity_type"))
        if not name:
            continue
        if name in workflow_names:
            msg = f"Duplicate workflow '{name}'."
            raise ValueError(msg)
        if entity_type and entity_type not in entity_type_names:
            msg = f"workflows row {index}: unknown entity_type '{entity_type}'."
            raise ValueError(msg)
        if not entity_type:
            msg = f"workflows row {index}: entity_type is required."
            raise ValueError(msg)
        workflow_names.add(name)
        workflows.append(
            {
                "name": name,
                "entity_type": entity_type,
                "auto_first_step": _truthy(_cell_str(item.get("auto_first_step"))),
                "publish": _truthy(_cell_str(item.get("publish") or "true")),
            }
        )

    workflow_followups: list[dict[str, Any]] = []
    for index, item in enumerate(data.get("workflow_followups") or [], start=1):
        workflow_name = _cell_str(item.get("workflow"))
        form_name = _cell_str(item.get("form"))
        if not workflow_name and not form_name:
            continue
        if workflow_name not in workflow_names:
            msg = f"workflow_followups row {index}: unknown workflow '{workflow_name}'."
            raise ValueError(msg)
        if form_name not in form_names:
            msg = f"workflow_followups row {index}: unknown form '{form_name}'."
            raise ValueError(msg)
        order_raw = item.get("order")
        try:
            order = int(order_raw) if order_raw not in (None, "") else 0
        except (TypeError, ValueError) as exc:
            msg = f"workflow_followups row {index}: order must be an integer."
            raise ValueError(msg) from exc
        if order < 0:
            msg = f"workflow_followups row {index}: order must be >= 0."
            raise ValueError(msg)
        try:
            condition = parse_condition(item.get("condition"))
        except json.JSONDecodeError as exc:
            msg = f"workflow_followups row {index}: condition is not valid JSON."
            raise ValueError(msg) from exc
        workflow_followups.append(
            {
                "workflow": workflow_name,
                "form": form_name,
                "order": order,
                "condition": condition,
            }
        )

    workflow_changes: list[dict[str, Any]] = []
    for index, item in enumerate(data.get("workflow_changes") or [], start=1):
        workflow_name = _cell_str(item.get("workflow"))
        form_name = _cell_str(item.get("form"))
        if not workflow_name and not form_name:
            continue
        if workflow_name not in workflow_names:
            msg = f"workflow_changes row {index}: unknown workflow '{workflow_name}'."
            raise ValueError(msg)
        if form_name not in form_names:
            msg = f"workflow_changes row {index}: unknown form '{form_name}'."
            raise ValueError(msg)
        try:
            mapping = item.get("mapping")
            if isinstance(mapping, dict):
                mapping = {str(key): str(val) for key, val in mapping.items()}
            else:
                mapping = parse_mapping(mapping)
        except (json.JSONDecodeError, ValueError) as exc:
            msg = f"workflow_changes row {index}: invalid mapping ({exc})."
            raise ValueError(msg) from exc
        workflow_changes.append(
            {
                "workflow": workflow_name,
                "form": form_name,
                "mapping": mapping,
            }
        )

    if (workflow_followups or workflow_changes) and not workflows:
        msg = "workflow_followups / workflow_changes need a workflows sheet with at least one row."
        raise ValueError(msg)

    return {
        "settings": settings,
        "set_as_default": _truthy(settings["set_as_default"]),
        "org_unit_types": types,
        "org_units": rows,
        "org_unit_geo": org_unit_geo,
        "forms": forms,
        "entity_types": entity_types,
        "entities": entities,
        "instances": instances,
        "workflows": workflows,
        "workflow_followups": workflow_followups,
        "workflow_changes": workflow_changes,
    }


def unique_nodes(
    types: list[dict[str, Any]],
    rows: list[dict[str, str]],
) -> list[tuple[tuple[str, ...], str, int]]:
    """Return (path, name, type_index) in parent-before-child order."""
    seen: set[tuple[str, ...]] = set()
    nodes: list[tuple[tuple[str, ...], str, int]] = []
    for row in rows:
        path: list[str] = []
        for index, level in enumerate(types):
            name = row[level["name"]]
            path.append(name)
            key = tuple(path)
            if key not in seen:
                seen.add(key)
                nodes.append((key, name, index))
    return nodes


def write_workbook(path: Path, data: dict[str, Any] | None = None) -> Path:
    payload = validate_workbook(data or sample_workbook())
    settings = payload["settings"]
    types = payload["org_unit_types"]
    rows = payload["org_units"]
    type_names = [item["name"] for item in types]

    workbook = Workbook()
    workbook.properties.title = "IASO account setup"
    workbook.properties.creator = "iaso-scripts"
    workbook.properties.description = (
        "Template to fill a connected IASO account with demo pyramid, geography, forms, entities and workflows."
    )

    readme = workbook.active
    readme.title = SHEET_README
    readme.sheet_properties.tabColor = _TAB[SHEET_README]
    readme.sheet_view.showGridLines = False
    readme.column_dimensions["A"].width = 108
    for line in _README.splitlines():
        _append_row(readme, [line])
        row = readme.max_row
        cell = readme.cell(row, 1)
        cell.alignment = _WRAP_TOP
        text = line.strip()
        if row == 1:
            cell.font = _TITLE_FONT
            cell.fill = _TITLE_FILL
            readme.row_dimensions[row].height = 32
        elif text in {"How to run", "Sheets", "Rules"}:
            cell.font = _SECTION_FONT
            cell.fill = _SECTION_FILL
            readme.row_dimensions[row].height = 22
        elif text[:1].isdigit() and ". " in text[:4]:
            cell.font = _README_FONT
        elif text.startswith(("-", "uv ", "or MCP")):
            cell.font = _README_MUTED
        else:
            cell.font = _README_FONT
            if text:
                readme.row_dimensions[row].height = 18

    settings_sheet = workbook.create_sheet(SHEET_SETTINGS)
    _header_row(
        settings_sheet,
        ["key", "value", "notes"],
        comments={
            "key": "Do not rename these keys — the importer matches them.",
            "value": "Edit the yellow cells.",
            "notes": "Guidance only; ignored by the importer.",
        },
    )
    notes = {
        "project_name": "Matched on re-run; created if missing.",
        "project_app_id": "Must be unique on the IASO instance for a new project.",
        "project_description": "Optional.",
        "project_color": "Hex color, e.g. #1976D2. Used for the project in IASO.",
        "data_source_name": "Matched on re-run; created if missing.",
        "source_version_description": "Used when creating the first source version.",
        "set_as_default": "true/false. Sets this source as the account default version.",
    }
    for key in SETTING_KEYS:
        _append_row(settings_sheet, [key, settings[key], notes.get(key, "")])
    _style_data_rows(settings_sheet, skip_cols={1, 3})
    for row in range(2, settings_sheet.max_row + 1):
        settings_sheet.cell(row, 1).font = _KEY_FONT
        settings_sheet.cell(row, 1).fill = _KEY_FILL
        settings_sheet.cell(row, 2).fill = _EDIT_FILL
        settings_sheet.cell(row, 3).font = _NOTE_FONT
        settings_sheet.cell(row, 3).fill = _KEY_FILL
    color_value = settings.get("project_color") or ""
    if color_value.startswith("#") and len(color_value) == 7:
        try:
            hex_color = color_value[1:].upper()
            int(hex_color, 16)
            swatch = settings_sheet.cell(SETTING_KEYS.index("project_color") + 2, 2)
            swatch.fill = _fill(hex_color)
            swatch.font = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
        except ValueError:
            pass
    last_settings = settings_sheet.max_row
    _dropdown(
        settings_sheet,
        '"true,false"',
        f"B{SETTING_KEYS.index('set_as_default') + 2}",
        "true = use this pyramid as the account default (needed by the IASO UI).",
    )
    _autosize(settings_sheet, min_width=22, max_width=64)

    types_sheet = workbook.create_sheet(SHEET_TYPES)
    _header_row(
        types_sheet,
        ["name", "short_name", "depth", "parent_type"],
        comments={
            "name": "Shown in IASO. org_units column titles must match this exactly.",
            "short_name": "Short code, e.g. HF.",
            "depth": "1 = root (country). No gaps: 1, 2, 3…",
            "parent_type": "Leave empty on depth 1. Otherwise the name of the parent type.",
        },
    )
    for item in types:
        _append_row(
            types_sheet,
            [item["name"], item["short_name"], item["depth"], item["parent_type"]],
        )
    _style_data_rows(types_sheet)
    last_types = max(2, types_sheet.max_row)
    _dropdown(
        types_sheet,
        f"{SHEET_TYPES}!$A$2:$A${last_types}",
        f"D2:D{last_types}",
        "Parent type name (empty for the root).",
    )
    _autosize(types_sheet)

    org_sheet = workbook.create_sheet(SHEET_ORG_UNITS)
    header_fills = [_LEVEL_HEADER[index % len(_LEVEL_HEADER)] for index in range(len(type_names))]
    level_fills = [_LEVEL_CELL[index % len(_LEVEL_CELL)] for index in range(len(type_names))]
    _header_row(
        org_sheet,
        type_names,
        fills=header_fills,
        comments={
            name: (
                "Root of the pyramid. Same name is reused across rows."
                if index == 0
                else "Leaf facility — one row per site."
                if index == len(type_names) - 1
                else "Parent of the column to the right. Unique names help."
            )
            for index, name in enumerate(type_names)
        },
    )
    for row in rows:
        _append_row(org_sheet, [row[name] for name in type_names])
    _style_data_rows(org_sheet, level_fills=level_fills)
    last_org = max(2, org_sheet.max_row)
    last_org_col = get_column_letter(len(type_names) or 1)
    _autosize(org_sheet, min_width=16, max_width=40)

    geo_sheet = workbook.create_sheet(SHEET_GEO)
    _header_row(
        geo_sheet,
        ["name", "latitude", "longitude", "altitude", "shape"],
        comments={
            "name": "Must match an org unit name on org_units (country, region, district, or facility).",
            "latitude": "GPS point. WGS84. Empty on admin units that only have a shape.",
            "longitude": "GPS point. WGS84. Pair with latitude.",
            "altitude": "Optional metres. Ignored if there is no point.",
            "shape": (
                "Polygon: GeoJSON, or bbox west,south,east,north "
                "(min lon, min lat, max lon, max lat). Nested boxes work well for demos."
            ),
        },
    )
    for item in payload.get("org_unit_geo") or []:
        lat = item.get("latitude")
        lng = item.get("longitude")
        alt = item.get("altitude")
        _append_row(
            geo_sheet,
            [
                item["name"],
                "" if lat is None else lat,
                "" if lng is None else lng,
                "" if alt is None else alt,
                shape_to_cell(item.get("shape")),
            ],
        )
    _style_data_rows(geo_sheet)
    last_geo = max(2, geo_sheet.max_row)
    _dropdown(
        geo_sheet,
        f"{SHEET_ORG_UNITS}!$A$2:${last_org_col}${last_org}",
        f"A2:A{last_geo}",
        "Org unit name from the pyramid (any column).",
    )
    _autosize(geo_sheet, min_width=14, max_width=56)

    forms_sheet = workbook.create_sheet(SHEET_FORMS)
    _header_row(
        forms_sheet,
        ["name", "form_id", "org_unit_type", "kind"],
        comments={
            "name": "Shown in IASO. entity_types.reference_form and instances.form use this.",
            "form_id": "ODK form id (letters, numbers, underscore). Must stay unique.",
            "org_unit_type": "Must match org_unit_types.name (dropdown).",
            "kind": "profile = entity attributes. visit = follow-up submissions. update = profile edits.",
        },
    )
    for item in payload["forms"]:
        _append_row(
            forms_sheet,
            [item["name"], item["form_id"], item["org_unit_type"], item["kind"]],
        )
    _style_data_rows(forms_sheet)
    last_forms = max(2, forms_sheet.max_row)
    _dropdown(
        forms_sheet,
        f"{SHEET_TYPES}!$A$2:$A${last_types}",
        f"C2:C{last_forms}",
        "Org unit type this form can be filled on.",
    )
    _dropdown(
        forms_sheet,
        '"profile,visit,update,form"',
        f"D2:D{last_forms}",
        "profile = beneficiary file. visit = later submissions. update = edit the file.",
    )
    _autosize(forms_sheet)

    et_sheet = workbook.create_sheet(SHEET_ENTITY_TYPES)
    _header_row(
        et_sheet,
        [
            "name",
            "reference_form",
            "list_fields",
            "detail_fields",
            "duplicate_fields",
        ],
        comments={
            "name": "e.g. Beneficiary. entities.entity_type must match this.",
            "reference_form": "Must match forms.name of a profile form (dropdown).",
            "list_fields": "Comma-separated question names shown in the entity list.",
            "detail_fields": "Comma-separated question names on the entity detail page.",
            "duplicate_fields": "Used to search for possible duplicates.",
        },
    )
    for item in payload["entity_types"]:
        _append_row(
            et_sheet,
            [
                item["name"],
                item["reference_form"],
                _join_fields(item["list_fields"]),
                _join_fields(item["detail_fields"]),
                _join_fields(item["duplicate_fields"]),
            ],
        )
    _style_data_rows(et_sheet)
    last_et = max(2, et_sheet.max_row)
    _dropdown(
        et_sheet,
        f"{SHEET_FORMS}!$A$2:$A${last_forms}",
        f"B2:B{last_et}",
        "Profile form that holds core attributes.",
    )
    _autosize(et_sheet, min_width=16, max_width=48)

    wf_sheet = workbook.create_sheet(SHEET_WORKFLOWS)
    _header_row(
        wf_sheet,
        ["name", "entity_type", "auto_first_step", "publish"],
        comments={
            "name": "Workflow version name, e.g. Beneficiary care.",
            "entity_type": "Must match entity_types.name (dropdown). One workflow per type.",
            "auto_first_step": "true = mobile opens the first follow-up right after registration.",
            "publish": "true = publish after creating follow-ups and changes (needed on mobile).",
        },
    )
    for item in payload["workflows"]:
        _append_row(
            wf_sheet,
            [
                item["name"],
                item["entity_type"],
                "true" if item.get("auto_first_step") else "false",
                "true" if item.get("publish") else "false",
            ],
        )
    _style_data_rows(wf_sheet)
    last_wf = max(2, wf_sheet.max_row)
    _dropdown(
        wf_sheet,
        f"{SHEET_ENTITY_TYPES}!$A$2:$A${last_et}",
        f"B2:B{last_wf}",
        "Entity type this workflow belongs to.",
    )
    _dropdown(
        wf_sheet,
        '"true,false"',
        f"C2:C{last_wf}",
        "Jump into the first follow-up after registration.",
    )
    _dropdown(
        wf_sheet,
        '"true,false"',
        f"D2:D{last_wf}",
        "Publish so the mobile app can use this version.",
    )
    _autosize(wf_sheet)

    fu_sheet = workbook.create_sheet(SHEET_FOLLOWUPS)
    _header_row(
        fu_sheet,
        ["workflow", "form", "order", "condition"],
        comments={
            "workflow": "Must match workflows.name (dropdown).",
            "form": "Follow-up form shown on mobile. Must match forms.name.",
            "order": "0 is first. auto_first_step uses the first follow-up.",
            "condition": 'true = always. Or JsonLogic JSON, e.g. {"==":[{"var":"sex"},"F"]}.',
        },
    )
    for item in payload["workflow_followups"]:
        condition = item.get("condition")
        condition_cell = json.dumps(condition, separators=(",", ":")) if not isinstance(condition, str) else condition
        _append_row(
            fu_sheet,
            [item["workflow"], item["form"], item.get("order", 0), condition_cell],
        )
    _style_data_rows(fu_sheet)
    last_fu = max(2, fu_sheet.max_row)
    _dropdown(
        fu_sheet,
        f"{SHEET_WORKFLOWS}!$A$2:$A${last_wf}",
        f"A2:A{last_fu}",
        "Workflow this follow-up belongs to.",
    )
    _dropdown(
        fu_sheet,
        f"{SHEET_FORMS}!$A$2:$A${last_forms}",
        f"B2:B{last_fu}",
        "Form offered after registration.",
    )
    _autosize(fu_sheet)

    ch_sheet = workbook.create_sheet(SHEET_CHANGES)
    _header_row(
        ch_sheet,
        ["workflow", "form", "mapping"],
        comments={
            "workflow": "Must match workflows.name (dropdown).",
            "form": "Source form (not the profile). Answers copy onto the reference form.",
            "mapping": "first_name,last_name or first_name=first_name. Empty = overlapping questions.",
        },
    )
    for item in payload["workflow_changes"]:
        _append_row(
            ch_sheet,
            [item["workflow"], item["form"], _join_mapping(item.get("mapping"))],
        )
    _style_data_rows(ch_sheet)
    last_ch = max(2, ch_sheet.max_row)
    _dropdown(
        ch_sheet,
        f"{SHEET_WORKFLOWS}!$A$2:$A${last_wf}",
        f"A2:A{last_ch}",
        "Workflow this change belongs to.",
    )
    _dropdown(
        ch_sheet,
        f"{SHEET_FORMS}!$A$2:$A${last_forms}",
        f"B2:B{last_ch}",
        "Update/visit form whose answers write back to the profile.",
    )
    _autosize(ch_sheet)

    entity_headers = ["name", "entity_type", "org_unit"]
    for row in payload["entities"]:
        for key in row:
            if key not in entity_headers:
                entity_headers.append(key)
    entities_sheet = workbook.create_sheet(SHEET_ENTITIES)
    _header_row(
        entities_sheet,
        entity_headers,
        comments={
            "name": "Display name. Keep it unique. Used to build a stable uuid.",
            "entity_type": "Must match entity_types.name (dropdown).",
            "org_unit": "Must match a facility name on org_units (dropdown on the leaf column).",
            "first_name": "Extra columns become answers on the profile form.",
            "last_name": "Extra columns become answers on the profile form.",
            "sex": "F or M (dropdown).",
            "age": "Whole number.",
        },
    )
    for row in payload["entities"]:
        _append_row(entities_sheet, [row.get(key, "") for key in entity_headers])
    _style_data_rows(entities_sheet)
    last_ent = max(2, entities_sheet.max_row)
    _dropdown(
        entities_sheet,
        f"{SHEET_ENTITY_TYPES}!$A$2:$A${last_et}",
        f"B2:B{last_ent}",
        "Entity type.",
    )
    _dropdown(
        entities_sheet,
        f"{SHEET_ORG_UNITS}!${last_org_col}$2:${last_org_col}${last_org}",
        f"C2:C{last_ent}",
        "Health facility (leaf org unit).",
    )
    if "sex" in entity_headers:
        sex_col = get_column_letter(entity_headers.index("sex") + 1)
        _dropdown(
            entities_sheet,
            '"F,M"',
            f"{sex_col}2:{sex_col}{last_ent}",
            "F = female, M = male.",
        )
    _autosize(entities_sheet)

    instance_headers = ["entity", "form", "org_unit"]
    for row in payload["instances"]:
        for key in row:
            if key not in instance_headers:
                instance_headers.append(key)
    instances_sheet = workbook.create_sheet(SHEET_INSTANCES)
    _header_row(
        instances_sheet,
        instance_headers,
        comments={
            "entity": "Must match entities.name (dropdown).",
            "form": "Must match forms.name of a visit form (dropdown).",
            "org_unit": "Facility where the visit happened (leaf org unit).",
            "visit_date": "ISO date, e.g. 2026-03-12.",
        },
    )
    for row in payload["instances"]:
        _append_row(instances_sheet, [row.get(key, "") for key in instance_headers])
    _style_data_rows(instances_sheet)
    last_inst = max(2, instances_sheet.max_row)
    _dropdown(
        instances_sheet,
        f"{SHEET_ENTITIES}!$A$2:$A${last_ent}",
        f"A2:A{last_inst}",
        "Beneficiary this visit belongs to.",
    )
    _dropdown(
        instances_sheet,
        f"{SHEET_FORMS}!$A$2:$A${last_forms}",
        f"B2:B{last_inst}",
        "Visit form.",
    )
    _dropdown(
        instances_sheet,
        f"{SHEET_ORG_UNITS}!${last_org_col}$2:${last_org_col}${last_org}",
        f"C2:C{last_inst}",
        "Health facility.",
    )
    _autosize(instances_sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _read_table(workbook, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name not in workbook.sheetnames:
        return []
    rows_iter = workbook[sheet_name].iter_rows(values_only=True)
    header = [_cell_str(value) for value in next(rows_iter, ())]
    records: list[dict[str, str]] = []
    for row in rows_iter:
        record = {
            header[index]: _cell_str(row[index] if index < len(row) else "")
            for index, name in enumerate(header)
            if name
        }
        if any(record.values()):
            records.append(record)
    return records


def write_sample_template(path: Path | None = None) -> Path:
    target = path or DEFAULT_XLSX_PATH
    return write_workbook(target, sample_workbook())


def read_workbook(path: Path) -> dict[str, Any]:
    if not path.is_file():
        msg = f"Workbook not found: {path}"
        raise FileNotFoundError(msg)
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if SHEET_SETTINGS not in workbook.sheetnames:
            msg = f"Missing sheet '{SHEET_SETTINGS}'. Found: {workbook.sheetnames}."
            raise ValueError(msg)
        if SHEET_TYPES not in workbook.sheetnames:
            msg = f"Missing sheet '{SHEET_TYPES}'. Found: {workbook.sheetnames}."
            raise ValueError(msg)
        if SHEET_ORG_UNITS not in workbook.sheetnames:
            msg = f"Missing sheet '{SHEET_ORG_UNITS}'. Found: {workbook.sheetnames}."
            raise ValueError(msg)

        settings: dict[str, str] = {}
        sheet = workbook[SHEET_SETTINGS]
        rows = sheet.iter_rows(values_only=True)
        header = [_cell_str(value).lower() for value in next(rows, ())]
        try:
            key_i = header.index("key")
            value_i = header.index("value")
        except ValueError as exc:
            msg = "settings sheet needs columns key, value."
            raise ValueError(msg) from exc
        for row in rows:
            key = _cell_str(row[key_i] if key_i < len(row) else "")
            if key:
                settings[key] = _cell_str(row[value_i] if value_i < len(row) else "")

        types: list[dict[str, Any]] = []
        sheet = workbook[SHEET_TYPES]
        rows = sheet.iter_rows(values_only=True)
        header = [_cell_str(value) for value in next(rows, ())]
        header_l = [name.lower() for name in header]
        required = ("name", "short_name", "depth")
        missing = [name for name in required if name not in header_l]
        if missing:
            msg = f"org_unit_types missing columns {missing}."
            raise ValueError(msg)
        name_i = header_l.index("name")
        short_i = header_l.index("short_name")
        depth_i = header_l.index("depth")
        parent_i = header_l.index("parent_type") if "parent_type" in header_l else None
        for row in rows:
            name = _cell_str(row[name_i] if name_i < len(row) else "")
            if not name:
                continue
            types.append(
                {
                    "name": name,
                    "short_name": _cell_str(row[short_i] if short_i < len(row) else ""),
                    "depth": row[depth_i] if depth_i < len(row) else None,
                    "parent_type": (_cell_str(row[parent_i] if parent_i is not None and parent_i < len(row) else "")),
                }
            )

        sheet = workbook[SHEET_ORG_UNITS]
        rows_iter = sheet.iter_rows(values_only=True)
        ou_header = [_cell_str(value) for value in next(rows_iter, ())]
        org_units: list[dict[str, str]] = []
        for row in rows_iter:
            record = {
                ou_header[index]: _cell_str(row[index] if index < len(row) else "")
                for index, name in enumerate(ou_header)
                if name
            }
            if any(record.values()):
                org_units.append(record)

        forms = _read_table(workbook, SHEET_FORMS)
        entity_types = _read_table(workbook, SHEET_ENTITY_TYPES)
        entities = _read_table(workbook, SHEET_ENTITIES)
        instances = _read_table(workbook, SHEET_INSTANCES)
        workflows = _read_table(workbook, SHEET_WORKFLOWS)
        workflow_followups = _read_table(workbook, SHEET_FOLLOWUPS)
        workflow_changes = _read_table(workbook, SHEET_CHANGES)
        org_unit_geo = _read_table(workbook, SHEET_GEO)
    finally:
        workbook.close()

    payload = validate_workbook(
        {
            "settings": settings,
            "org_unit_types": types,
            "org_units": org_units,
            "org_unit_geo": org_unit_geo,
            "forms": forms,
            "entity_types": entity_types,
            "entities": entities,
            "instances": instances,
            "workflows": workflows,
            "workflow_followups": workflow_followups,
            "workflow_changes": workflow_changes,
        }
    )
    payload["xlsx"] = path.name
    payload["xlsx_path"] = str(path)
    return payload


def summarize_workbook(data: dict[str, Any]) -> dict[str, Any]:
    types = data["org_unit_types"]
    nodes = unique_nodes(types, data["org_units"])
    counts = []
    for index, level in enumerate(types):
        counts.append(
            {
                "depth": level["depth"],
                "name": level["name"],
                "short_name": level["short_name"],
                "count": sum(1 for _path, _name, type_index in nodes if type_index == index),
            }
        )
    sample = [
        {
            "path": " / ".join(path),
            "name": name,
            "type": types[type_index]["name"],
            "depth": types[type_index]["depth"],
        }
        for path, name, type_index in nodes[:25]
    ]
    settings = data["settings"]
    return {
        "xlsx": data.get("xlsx"),
        "xlsx_path": data.get("xlsx_path"),
        "project_name": settings["project_name"],
        "project_app_id": settings["project_app_id"],
        "data_source_name": settings["data_source_name"],
        "set_as_default": data["set_as_default"],
        "rows": len(data["org_units"]),
        "org_units": len(nodes),
        "geo_points": sum(1 for item in data.get("org_unit_geo") or [] if item.get("latitude") is not None),
        "geo_shapes": sum(1 for item in data.get("org_unit_geo") or [] if item.get("shape")),
        "types": [
            {
                "name": level["name"],
                "short_name": level["short_name"],
                "depth": level["depth"],
                "parent_type": level["parent_type"],
            }
            for level in types
        ],
        "counts_by_level": counts,
        "sample_org_units": sample,
        "forms": [
            {
                "name": item["name"],
                "form_id": item["form_id"],
                "kind": item["kind"],
                "org_unit_type": item["org_unit_type"],
            }
            for item in data.get("forms") or []
        ],
        "entity_types": [
            {"name": item["name"], "reference_form": item["reference_form"]} for item in data.get("entity_types") or []
        ],
        "entities_count": len(data.get("entities") or []),
        "instances_count": len(data.get("instances") or []),
        "workflows": [
            {
                "name": item["name"],
                "entity_type": item["entity_type"],
                "auto_first_step": item.get("auto_first_step"),
                "publish": item.get("publish"),
            }
            for item in data.get("workflows") or []
        ],
        "workflow_followups_count": len(data.get("workflow_followups") or []),
        "workflow_changes_count": len(data.get("workflow_changes") or []),
        "sample_entities": [
            {
                "name": row.get("name"),
                "org_unit": row.get("org_unit"),
                "entity_type": row.get("entity_type"),
            }
            for row in (data.get("entities") or [])[:10]
        ],
    }
