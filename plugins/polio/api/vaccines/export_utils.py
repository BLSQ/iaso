from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from plugins.polio.api.vaccines.common import sort_results
from plugins.polio.export_utils import cell_border


COMMON_COLUMNS = ["Country", "Vaccine", "Date", "Vials type", "Action Type", "Action"]
USABLE_COLUMNS = ["Vials IN", "Vials OUT", "Doses IN", "Doses OUT"]
UNUSABLE_COLUMNS = ["Vials IN", "Vials OUT"]
COMMON_DATA_KEYS = ["date", "vials_type", "type", "action"]
USABLE_DATA_KEYS = ["vials_in", "vials_out", "doses_in", "doses_out"]
UNUSABLE_DATA_KEYS = ["vials_in", "vials_out"]


def get_sheet_configs():
    data_keys = ["country", "vaccine"] + COMMON_DATA_KEYS

    variants_columns_keys = {
        "Usable": {
            "extra_columns": USABLE_COLUMNS,
            "extra_keys": USABLE_DATA_KEYS,
        },
        "Unusable": {
            "extra_columns": UNUSABLE_COLUMNS,
            "extra_keys": UNUSABLE_DATA_KEYS,
        },
        "Earmarked": {
            "extra_columns": UNUSABLE_COLUMNS,
            "extra_keys": UNUSABLE_DATA_KEYS,
        },
    }

    return {
        name: {
            "columns": COMMON_COLUMNS + variant["extra_columns"],
            "keys": data_keys + variant["extra_keys"],
            "sum_columns": variant["extra_keys"],
        }
        for name, variant in variants_columns_keys.items()
    }


def write_vials_doses_totals(sheet, config, sums, sum_columns_indices):
    """
    Writes the total row into the Excel sheet.
    """
    action_index = config["keys"].index("action")

    total_row = [""] * len(config["columns"])
    total_row[action_index] = "Total :"

    for col, idx in zip(config["sum_columns"], sum_columns_indices):
        total_row[idx] = sums[col]

    sheet.append(total_row)
    total_row_index = sheet.max_row
    for col in range(1, len(config["columns"]) + 1):
        cell = sheet.cell(row=total_row_index, column=col)
        cell.font = Font(bold=True)
        cell.number_format = "#,##0"


def vials_doses_totals(data, sum_columns):
    """
    Calculates the total sum for each column in sum_columns.
    """
    sums = dict.fromkeys(sum_columns, 0)
    for entry in data:
        for col in sum_columns:
            if isinstance(entry.get(col, 0), (int, float)):
                sums[col] += entry[col]
    return sums


def write_vials_doses_stock_balance(sheet, config, sums, sum_columns_indices):
    """
    Adds stock balance row to the Excel sheet.
    """
    action_index = config["keys"].index("action")
    stock_balances_row = [""] * len(config["columns"])
    stock_balances_row[action_index] = "Stock Balances :"

    if "vials_in" in sums and "vials_out" in sums:
        stock_balances_row[sum_columns_indices[0]] = sums["vials_in"] - sums["vials_out"]

    if "doses_in" in sums and "doses_out" in sums:
        stock_balances_row[sum_columns_indices[2]] = sums["doses_in"] - sums["doses_out"]

    sheet.append(stock_balances_row)

    stock_balances_row_index = sheet.max_row
    for col in range(1, len(config["columns"]) + 1):
        cell = sheet.cell(row=stock_balances_row_index, column=col)
        cell.font = Font(bold=True)
        cell.number_format = "#,##0"


def write_colums_headers(sheet, config_columns):
    """
    Adds columns headers.
    """
    columns = config_columns["columns"] if isinstance(config_columns, dict) else config_columns
    sheet.append(columns)
    for col, _ in enumerate(columns, start=1):
        sheet.column_dimensions[chr(64 + col)].width = 21
        cell_header = sheet.cell(row=1, column=col)
        cell_header = cell_border(cell_header)


def write_columns_data(sheet, config_keys, datas, sheet_name):
    """
    Adds columns data.
    """

    data_keys = config_keys["keys"] if isinstance(config_keys, dict) else config_keys
    for entry in datas:
        row = []
        for key in data_keys:
            if key == "vials_type":
                row.append(sheet_name)
            else:
                row.append(entry[key] if entry[key] is not None else "")
        sheet.append(row)

        for col_index, value in enumerate(row, start=1):
            if isinstance(value, (int, float)):
                sheet.cell(row=sheet.max_row, column=col_index).number_format = "#,##0"


def get_vaccine_country(vaccine_stock):
    country = vaccine_stock.country.name
    vaccine = vaccine_stock.vaccine
    return {"country": country, "vaccine": vaccine}


def download_xlsx_stock_variants(request, filename, results, lambda_methods, vaccince_stock, tab):
    workbook = Workbook()
    sheet_configs = get_sheet_configs()

    sheets_order = sheet_configs.keys()
    sheets = {}
    for sheet_name in sheets_order:
        config = sheet_configs[sheet_name]
        if sheet_name == tab:
            sheet = workbook.active
            sheet.title = sheet_name
        else:
            sheet = workbook.create_sheet(sheet_name)

        sheets[sheet_name] = sheet

        write_colums_headers(sheet, config)

        datas = results if sheet_name == tab else sort_results(request, lambda_methods.get(sheet_name, list)())
        vaccine_country = get_vaccine_country(vaccince_stock)
        for data in datas:
            data["country"] = vaccine_country["country"]
            data["vaccine"] = vaccine_country["vaccine"]

        write_columns_data(sheet, config, datas, sheet_name)

        sheet.append([""] * len(config["columns"]))

        sum_columns_indices = [config["keys"].index(col) for col in config["sum_columns"]]
        sums = vials_doses_totals(datas, config["sum_columns"])
        write_vials_doses_totals(sheet, config, sums, sum_columns_indices)

        write_vials_doses_stock_balance(sheet, config, sums, sum_columns_indices)

    workbook._sheets = [sheets[name] for name in sheets_order]
    workbook.save(filename)
    return workbook


def write_vials_doses_total(
    sheet,
    total_vials,
    total_doses,
    vials_in,
    vials_out,
    doses_in,
    doses_out,
    earmarked_totals=None,
):
    total_vials_in_label = "Total Vials IN"
    total_doses_in_label = "Total Doses IN"
    total_vials_out_label = "Total Vials OUT"
    total_doses_out_label = "Total Doses OUT"
    stock_balance_label = "Stock Balance"
    earmarked_label = "Earmarked Stock"

    total_labels = [
        "",
        "",
        "",
        "",
        "",
        "",
        total_vials_in_label,
        total_vials_out_label,
        total_doses_in_label,
        total_doses_out_label,
    ]
    sheet.append(total_labels)

    total_values = ["", "", "", "", "", "Total", vials_in, vials_out, doses_in, doses_out]
    sheet.append(total_values)
    # We only pass earmarked for usable vials
    if earmarked_totals:
        earmarked_values = ["", "", "", "", "", earmarked_label, "", earmarked_totals[0], "", earmarked_totals[1]]
        sheet.append(earmarked_values)
    stock_balance_values = ["", "", "", "", "", stock_balance_label, total_vials, "", total_doses, ""]
    sheet.append(stock_balance_values)

    total_labels_index = sheet.max_row - 3 if earmarked_totals else sheet.max_row - 2
    total_values_index = sheet.max_row - 2 if earmarked_totals else sheet.max_row - 1
    earmarked_values_index = sheet.max_row - 1
    stock_balance_values_index = sheet.max_row

    # Apply text styling
    for col in range(6, 11):
        labels = sheet.cell(row=total_labels_index, column=col)
        totals = sheet.cell(row=total_values_index, column=col)
        earmarked = sheet.cell(row=earmarked_values_index, column=col)
        stock_balance = sheet.cell(row=stock_balance_values_index, column=col)
        labels.font = Font(bold=True)
        totals.font = Font(bold=True)
        earmarked.font = Font(bold=True)
        stock_balance.font = Font(bold=True)
        labels.alignment = Alignment(horizontal="center")
        totals.alignment = Alignment(horizontal="center")
        earmarked.alignment = Alignment(horizontal="center")
        stock_balance.alignment = Alignment(horizontal="center")

    # format numbers
    for col in range(7, 11):
        cell = sheet.cell(row=total_values_index, column=col)
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

        cell = sheet.cell(row=earmarked_values_index, column=col)
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

        cell = sheet.cell(row=stock_balance_values_index, column=col)
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def download_xlsx_public_stock_variants(
    filename,
    usable_results,
    unusable_results,
    usable_totals,
    unusable_totals,
    earmarked_totals,
    usable_vials_in,
    usable_vials_out,
    usable_doses_in,
    usable_doses_out,
    unusable_vials_in,
    unusable_vials_out,
    unusable_doses_in,
    unusable_doses_out,
):
    workbook = Workbook()

    common_keys = ["country_name", "vaccine_type"] + COMMON_DATA_KEYS

    for idx, sheet_name in enumerate(["usable", "unusable"]):
        if idx == 0:
            sheet = workbook.active
            sheet.title = sheet_name
        else:
            sheet = workbook.create_sheet(sheet_name)
        columns = COMMON_COLUMNS + (USABLE_COLUMNS)

        write_colums_headers(sheet, columns)

        datas = usable_results if idx == 0 else unusable_results
        totals = usable_totals if idx == 0 else unusable_totals
        vials_in = usable_vials_in if idx == 0 else unusable_vials_in
        vials_out = usable_vials_out if idx == 0 else unusable_vials_out
        doses_in = usable_doses_in if idx == 0 else unusable_doses_in
        doses_out = usable_doses_out if idx == 0 else unusable_doses_out

        data_keys = common_keys + (USABLE_DATA_KEYS)
        write_columns_data(sheet, data_keys, datas, sheet_name)

        sheet.append([""] * len(columns))

        write_vials_doses_total(
            sheet=sheet,
            total_vials=totals[0],
            total_doses=totals[1],
            earmarked_totals=earmarked_totals if idx == 0 else None,
            vials_in=vials_in,
            vials_out=vials_out,
            doses_in=doses_in,
            doses_out=doses_out,
        )

    workbook.save(filename)
    return workbook
